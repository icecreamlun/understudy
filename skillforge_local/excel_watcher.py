"""Real local Excel watcher.

Watches one or more .xlsx files (or a directory of them) and turns saved edits
into normalized ``spreadsheet_row_updated`` activity events — the same event
shape Section A mines and the observation feed renders.

Mechanism: snapshot the workbook (per sheet, per row) with openpyxl, persist the
snapshot under .runtime/excel_watch, and on each change diff the new snapshot
against the last one. A simple mtime-polling loop drives it (stdlib only, works
everywhere); pass --once for a single pass (e.g. from a cron/hook).
"""

from __future__ import annotations

import hashlib
import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:  # openpyxl is already a project dependency
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _state_dir(root: Path) -> Path:
    d = root / ".runtime" / "excel_watch"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _state_path(root: Path, workbook: Path) -> Path:
    key = hashlib.sha1(str(workbook.resolve()).encode("utf-8")).hexdigest()[:16]
    return _state_dir(root) / f"{workbook.stem}-{key}.json"


def snapshot_workbook(path: Path) -> dict[str, dict[str, list[Any]]]:
    """Return {sheet_name: {row_number: [cell values]}} for the used range."""
    if load_workbook is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required for the Excel watcher (pip install openpyxl).")
    wb = load_workbook(path, read_only=True, data_only=True)
    snap: dict[str, dict[str, list[Any]]] = {}
    try:
        for ws in wb.worksheets:
            rows: dict[str, list[Any]] = {}
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = ["" if v is None else (v if isinstance(v, (str, int, float, bool)) else str(v)) for v in row]
                # Skip fully empty rows to keep snapshots compact.
                if any(v != "" for v in values):
                    rows[str(idx)] = values
            snap[ws.title] = rows
    finally:
        wb.close()
    return snap


def _headers(rows: dict[str, list[Any]]) -> list[str]:
    header_row = rows.get("1")
    if not header_row:
        return []
    return [str(v) for v in header_row]


def diff_snapshots(
    old: dict[str, dict[str, list[Any]]],
    new: dict[str, dict[str, list[Any]]],
) -> list[dict[str, Any]]:
    """Return a list of row-level changes between two snapshots."""
    changes: list[dict[str, Any]] = []
    for sheet, new_rows in new.items():
        old_rows = old.get(sheet, {})
        headers = _headers(new_rows) or _headers(old_rows)
        for row_no, new_values in new_rows.items():
            if row_no == "1":
                continue  # header row
            old_values = old_rows.get(row_no)
            if old_values == new_values:
                continue
            kind = "added" if old_values is None else "modified"
            cells: dict[str, Any] = {}
            width = max(len(new_values), len(old_values or []))
            for col in range(width):
                nv = new_values[col] if col < len(new_values) else ""
                ov = (old_values[col] if old_values and col < len(old_values) else "")
                if nv != ov:
                    label = headers[col] if col < len(headers) and headers[col] else f"Column {col + 1}"
                    cells[label] = nv
            changes.append({"sheet": sheet, "row_number": int(row_no), "kind": kind, "changes": cells})
    return changes


def events_from_changes(workbook: Path, changes: Iterable[dict[str, Any]], *, actor: str) -> list[dict[str, Any]]:
    ts = _utc_now()
    events = []
    for change in changes:
        events.append(
            {
                "event_id": f"excel_{workbook.stem}_{change['sheet']}_{change['row_number']}_{ts}",
                "type": "spreadsheet_row_updated",
                "ts": ts,
                "actor": actor,
                "payload": {
                    "workbook": str(workbook),
                    "sheet": change["sheet"],
                    "row_number": change["row_number"],
                    "kind": change["kind"],
                    "changes": change["changes"],
                },
            }
        )
    return events


def _append_events(events_log: Path, events: list[dict[str, Any]]) -> None:
    if not events:
        return
    events_log.parent.mkdir(parents=True, exist_ok=True)
    with events_log.open("a", encoding="utf-8") as f:
        for event in events:
            f.write(json.dumps(event, sort_keys=True) + "\n")


def scan_once(root: Path, workbook: Path, events_log: Path, *, actor: str = "excel_watch") -> list[dict[str, Any]]:
    """Snapshot the workbook, diff against the stored snapshot, append + return new events."""
    new_snap = snapshot_workbook(workbook)
    state_path = _state_path(root, workbook)
    old_snap: dict[str, Any] = {}
    if state_path.exists():
        try:
            old_snap = json.loads(state_path.read_text(encoding="utf-8"))
        except Exception:
            old_snap = {}
    changes = diff_snapshots(old_snap, new_snap) if old_snap else []
    events = events_from_changes(workbook, changes, actor=actor)
    _append_events(events_log, events)
    state_path.write_text(json.dumps(new_snap, sort_keys=True), encoding="utf-8")
    return events


def _resolve_workbooks(path: Path) -> list[Path]:
    if path.is_dir():
        return sorted(p for p in path.glob("*.xlsx") if not p.name.startswith("~$"))
    if path.suffix == ".xlsx":
        return [path]
    return []


def watch(
    root: Path,
    path: Path,
    events_log: Path,
    *,
    interval: float = 2.0,
    actor: str = "excel_watch",
    once: bool = False,
    on_event=None,
) -> int:
    """Poll the target path; emit events when watched workbooks change.

    Returns the number of events emitted (useful for --once)."""
    workbooks = _resolve_workbooks(path)
    if not workbooks:
        print(f"No .xlsx workbooks found at {path}")
        return 0

    mtimes: dict[Path, float] = {}
    total = 0

    def pass_once() -> int:
        nonlocal total
        emitted = 0
        for wb in workbooks:
            if not wb.exists():
                continue
            mtime = wb.stat().st_mtime
            if not once and mtimes.get(wb) == mtime:
                continue
            mtimes[wb] = mtime
            events = scan_once(root, wb, events_log, actor=actor)
            emitted += len(events)
            for event in events:
                payload = event["payload"]
                print(f"[excel-watch] {wb.name} · {payload['sheet']} row {payload['row_number']} {payload['kind']}")
                if on_event:
                    on_event(event)
        total += emitted
        return emitted

    if once:
        pass_once()
        return total

    print(f"[excel-watch] watching {len(workbooks)} workbook(s) under {path} (interval {interval}s). Ctrl-C to stop.")
    # First pass establishes baselines silently for files with no prior snapshot.
    pass_once()
    try:
        while True:
            time.sleep(interval)
            pass_once()
    except KeyboardInterrupt:
        print(f"\n[excel-watch] stopped after emitting {total} event(s).")
    return total
