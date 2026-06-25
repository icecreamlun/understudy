from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook


HEADERS = [
    "Received Date",
    "Customer",
    "Contact",
    "Request Type",
    "Due Date",
    "Blocker",
    "Owner",
    "Next Step",
    "Status",
    "Source Email",
    "Thread ID",
]


def build_row_preview(email_event: dict, extraction: dict) -> dict:
    payload = email_event["payload"]
    extracted = extraction["extracted"]
    return {
        "Received Date": email_event["ts"],
        "Customer": extracted["customer"],
        "Contact": extracted.get("contact") or payload.get("from", ""),
        "Request Type": extracted["request_type"],
        "Due Date": extracted.get("due_date", ""),
        "Blocker": "; ".join(extracted.get("blockers", [])),
        "Owner": "Data engineer",
        "Next Step": extracted.get("next_step", "Review request"),
        "Status": "Waiting on customer",
        "Source Email": payload["message_id"],
        "Thread ID": payload["thread_id"],
    }


def append_onboarding_row(workbook_path: Path, row: dict) -> int:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Onboarding Tracker"]
    headers = [cell.value for cell in sheet[1]]
    if headers != HEADERS:
        raise ValueError("Workbook headers do not match onboarding tracker schema")

    _raise_if_duplicate(sheet, row["Source Email"], row["Thread ID"])
    sheet.append([row.get(header, "") for header in HEADERS])
    row_number = sheet.max_row
    workbook.save(workbook_path)
    return row_number


def create_onboarding_tracker(workbook_path: Path) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Onboarding Tracker"
    sheet.append(HEADERS)
    workbook.save(workbook_path)


def ensure_onboarding_tracker(workbook_path: Path) -> None:
    if workbook_path.exists():
        return
    create_onboarding_tracker(workbook_path)


def _raise_if_duplicate(sheet, message_id: str, thread_id: str) -> None:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_message_id = row[9]
        existing_thread_id = row[10]
        if existing_message_id == message_id or existing_thread_id == thread_id:
            raise ValueError("Duplicate message_id or thread_id")
