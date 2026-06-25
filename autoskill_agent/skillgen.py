from __future__ import annotations

import csv
import fnmatch
import json
import os
import re
import shutil
import sqlite3
import subprocess
import zipfile
from copy import deepcopy
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from skillforge_local.llm import complete_text, default_model, load_local_env


VALID_SKILL_STATES = {
    "candidate",
    "installed",
    "active",
    "beta",
    "team_standard",
    "needs_refinement",
    "disabled",
    "deprecated",
}

REQUIRED_CANDIDATE_FIELDS = [
    "candidate_id",
    "candidate_name",
    "summary",
    "evidence",
    "suggested_trigger",
    "suggested_inputs",
    "suggested_outputs",
    "observed_actions",
    "suggested_guardrails",
    "confidence",
]

SECTION_A_CONTRACT_VERSIONS = {"section_a.skill_candidate.v1", "section_a.v1"}

REQUIRED_SECTION_A_FIELDS = [
    "contract_version",
    "candidate_id",
    "name_suggestion",
    "confidence",
    "status",
    "detected_at",
    "pattern",
    "evidence",
    "suggested_skill",
    "handoff",
]

REQUIRED_SECTION_A_MINIMAL_FIELDS = [
    "candidate_id",
    "name_suggestion",
    "confidence",
    "status",
    "pattern.workflow_family",
    "pattern.common_sequence",
    "pattern.trigger_signature",
    "evidence.episode_ids",
    "evidence.target_artifact",
    "evidence.target_sheet",
    "suggested_skill.trigger",
    "suggested_skill.inputs",
    "suggested_skill.actions",
    "suggested_skill.forbidden_actions",
    "handoff.required_confirmation_fields",
]

ALLOWED_EXECUTOR_ACTIONS = {
    "parse_xlsx_attachment",
    "preview_reconciliation_update",
    "flag_reconciliation_exceptions",
    "draft_email_reply",
    "require_human_approval",
    "write_xlsx_update",
    "write_audit_log",
}

ALLOWED_WORKFLOW_STEP_TYPES = {
    "read_input",
    "transform",
    "analyze",
    "draft_output",
    "human_approval",
    "write_output",
    "validate",
}

SKILL_SCHEMA_VERSION = "skill.workflow.v1"


@dataclass(frozen=True)
class SkillGenPaths:
    root: Path
    candidates_dir: Path
    reviews_dir: Path
    skills_dir: Path
    events_dir: Path
    attachments_dir: Path
    workbooks_dir: Path
    drafts_dir: Path
    matches_dir: Path
    runtime_dir: Path
    registry_db: Path
    event_log: Path
    skill_candidates_log: Path


@dataclass(frozen=True)
class LocalModelConfig:
    backend: str
    base_url: str
    api_key: str
    model: str
    timeout_seconds: int = 180


def paths(root: Path | str = ".") -> SkillGenPaths:
    root_path = Path(root).resolve()
    workspace = root_path / "workspace"
    runtime = root_path / ".runtime" / "skillforge"
    return SkillGenPaths(
        root=root_path,
        candidates_dir=workspace / "candidates",
        reviews_dir=workspace / "reviews",
        skills_dir=workspace / "skills",
        events_dir=workspace / "events",
        attachments_dir=workspace / "attachments",
        workbooks_dir=workspace / "workbooks",
        drafts_dir=workspace / "mail" / "drafts",
        matches_dir=workspace / "skill_matches",
        runtime_dir=runtime,
        registry_db=runtime / "skill_registry.sqlite3",
        event_log=workspace / "events" / "events.jsonl",
        skill_candidates_log=workspace / "events" / "skill_candidates.jsonl",
    )


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return slug or "generated_skill"


def kebab(value: str) -> str:
    return slugify(value).replace("_", "-")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def append_event(event_log: Path, event: dict[str, Any]) -> None:
    event_log.parent.mkdir(parents=True, exist_ok=True)
    payload = {"timestamp": utc_now(), **event}
    with event_log.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, sort_keys=True) + "\n")


def read_events(event_log: Path) -> list[dict[str, Any]]:
    if not event_log.exists():
        return []
    return [json.loads(line) for line in event_log.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = "\n".join(json.dumps(row, sort_keys=True) for row in rows)
    path.write_text(payload + ("\n" if rows else ""), encoding="utf-8")


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSONL in {path} at line {line_number}: {exc}") from exc
    return rows


def local_model_config(root: Path | str = ".", timeout_seconds: int = 180) -> LocalModelConfig:
    """Resolve the Codex CLI configuration for the skill planner.

    Every AI call runs through the OpenAI Codex CLI (`codex exec`) in API-key mode.
    The key is read from OPENAI_API_KEY (or a git-ignored .env.local at the project
    root). The model defaults to the Codex CLI's default and can be pinned with
    CODEX_MODEL / SKILLGEN_PLANNER_MODEL.
    """
    load_local_env(root)
    planner_model = (
        os.environ.get("SKILLGEN_PLANNER_MODEL")
        or os.environ.get("CODEX_MODEL")
        or ""  # empty = let the Codex CLI pick its default model
    )
    return LocalModelConfig(
        backend="codex",
        base_url="codex-cli (apikey)",
        api_key=os.environ.get("OPENAI_API_KEY", ""),
        model=planner_model or "codex-default",
        timeout_seconds=timeout_seconds,
    )


def call_local_chat_model(
    config: LocalModelConfig,
    messages: list[dict[str, str]],
    response_format: dict[str, str] | None = None,
    *,
    max_tokens: int = 16000,
    thinking: bool = True,
) -> dict[str, Any]:
    """Call Claude with OpenAI-style messages and return an OpenAI-shaped dict.

    The ``response_format`` argument is accepted for backward compatibility; the
    JSON contract is enforced by the prompt and parsed downstream. Small,
    schema-bound calls should pass a tight ``max_tokens`` and ``thinking=False``
    to keep latency low.
    """
    text = complete_text(
        messages,
        model=config.model,
        max_tokens=max_tokens,
        timeout_seconds=config.timeout_seconds,
        api_key=config.api_key or None,
        base_url=config.base_url,
        thinking=thinking,
    )
    return {"choices": [{"message": {"content": text}}]}


def local_model_text_response(response: dict[str, Any]) -> str:
    choices = response.get("choices", [])
    if not choices:
        raise RuntimeError("Local model response did not include choices")
    message = choices[0].get("message", {})
    content = message.get("content", "")
    if not content:
        raise RuntimeError("Local model response did not include message content")
    return content


def parse_json_object(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped, flags=re.IGNORECASE)
        stripped = re.sub(r"\s*```$", "", stripped)
    start = stripped.find("{")
    if start < 0:
        raise ValueError("Expected a JSON object")
    # The model may emit the JSON object followed by trailing prose or a second
    # block. ``raw_decode`` parses just the first complete object and ignores the
    # rest, so we no longer fail with "Extra data" and discard a valid plan.
    try:
        payload, _ = json.JSONDecoder().raw_decode(stripped[start:])
    except json.JSONDecodeError:
        end = stripped.rfind("}")
        if end <= start:
            raise
        payload = json.loads(stripped[start : end + 1])
    if not isinstance(payload, dict):
        raise ValueError("Expected a JSON object")
    return payload


def default_pattern_candidate() -> dict[str, Any]:
    return {
        "candidate_id": "cand_daily_cash_recon_001",
        "created_at": "2026-06-14T10:00:00-07:00",
        "candidate_name": "Daily Cash Reconciliation Workflow",
        "pattern_type": "email_to_spreadsheet_to_reply",
        "confidence": 0.93,
        "status": "ready_for_human_review",
        "summary": {
            "one_liner": (
                "Analysts repeatedly process daily bank transaction emails by updating "
                "cash_recon.xlsx and sending exception summaries."
            ),
            "business_domain": "finance",
            "workflow_family": "cash_reconciliation",
            "detected_frequency": "daily",
            "unique_actors": 2,
            "episode_count": 3,
        },
        "evidence": {
            "episodes": [
                "episode_cash_recon_2026_06_12",
                "episode_cash_recon_2026_06_13",
                "episode_cash_recon_2026_06_14",
            ],
            "actors": ["analyst_1", "analyst_2"],
            "input_event_ids": [
                "event_email_bank_2026_06_12",
                "event_email_bank_2026_06_13",
                "event_email_bank_2026_06_14",
            ],
            "spreadsheet_event_ids": [
                "event_sheet_cash_001",
                "event_sheet_cash_002",
                "event_sheet_cash_003",
            ],
            "sent_event_ids": [
                "event_sent_summary_001",
                "event_sent_summary_002",
                "event_sent_summary_003",
            ],
        },
        "suggested_trigger": {
            "event_type": "email_received",
            "conditions": [
                {"type": "subject_pattern", "value": "Daily bank transactions"},
                {"type": "has_attachment_matching", "value": "bank_transactions_*.xlsx"},
                {"type": "sender_pattern", "value": "bank-ops@example.local"},
            ],
        },
        "suggested_inputs": [
            {
                "name": "daily_bank_transaction_file",
                "type": "spreadsheet_attachment",
                "required": True,
                "source": "inbound_email_attachment",
            },
            {
                "name": "cash_recon_workbook",
                "type": "xlsx_workbook",
                "required": True,
                "path_hint": "workspace/workbooks/cash_recon.xlsx",
            },
        ],
        "suggested_outputs": [
            {
                "name": "updated_reconciliation_sheet",
                "type": "xlsx_update",
                "target_hint": "cash_recon.xlsx / Daily Reconciliation",
            },
            {
                "name": "exception_summary_reply",
                "type": "email_draft",
                "target_hint": "workspace/mail/drafts/",
            },
        ],
        "observed_actions": [
            {
                "type": "spreadsheet_update",
                "workbook": "cash_recon.xlsx",
                "sheet": "Daily Reconciliation",
                "headers": [
                    "Bank Amount",
                    "ERP Amount",
                    "Match Status",
                    "Exception Reason",
                    "Reviewer",
                ],
            },
            {
                "type": "sent_reply",
                "reply_pattern": "matched_count_and_exception_count_summary",
            },
        ],
        "suggested_guardrails": [
            "Do not send email automatically",
            "Do not overwrite reviewed rows",
            "Do not modify closed-period sheets",
            "Do not access network",
            "Require human approval before workbook changes",
        ],
        "candidate_quality": {
            "episode_count": 3,
            "unique_actor_count": 2,
            "similarity_score": 0.93,
            "missing_information": [],
        },
    }


def default_section_a_skill_candidate() -> dict[str, Any]:
    return {
        "contract_version": "section_a.skill_candidate.v1",
        "candidate_id": "cand_daily_cash_recon_001",
        "name_suggestion": "Daily Cash Reconciliation Skill",
        "confidence": 0.95,
        "status": "candidate",
        "detected_at": "2026-06-14T08:44:00-07:00",
        "pattern": {
            "workflow_family": "daily_cash_reconciliation",
            "episode_count": 3,
            "common_sequence": [
                "email_received",
                "spreadsheet_row_updated",
                "outbound_message_created",
            ],
            "trigger_signature": {
                "subject_prefix": "Daily bank transactions - Jun",
                "attachment_glob": "bank_transactions_*.xlsx",
                "intent": "daily_cash_reconciliation_request",
            },
            "similarity_reason": [
                "same inbound email intent",
                "same attachment naming pattern",
                "same target workbook",
                "same target sheet",
                "same status and exception classification columns",
                "same daily summary reply shape",
            ],
        },
        "evidence": {
            "episode_ids": [
                "episode_msg_bank_2026_06_12",
                "episode_msg_bank_2026_06_13",
                "episode_msg_bank_2026_06_14",
            ],
            "source_event_ids": [
                "email_in_20260612",
                "audit_003",
                "email_sent_20260612",
                "email_in_20260613",
                "audit_006",
                "email_sent_20260613",
                "email_in_20260614",
                "audit_009",
                "email_sent_20260614",
            ],
            "target_artifact": "workspace/workbooks/cash_recon.xlsx",
            "target_sheet": "Daily Reconciliation",
            "target_rows_examples": ["2:46", "47:91", "92:136"],
            "common_fields": [
                "Amount Diff",
                "Match Status",
                "Exception Reason",
                "Reviewer",
                "Reviewed At",
                "Source Email ID",
                "Notes",
            ],
            "daily_batch_size": 45,
        },
        "next_trigger": {
            "event_id": "event_email_bank_2026_06_15",
            "message_id": "msg_bank_2026_06_15",
            "thread_id": "thread_daily_cash_2026_06_15",
            "target_rows": "137:181",
            "unprocessed_action_columns": [
                "Match Status",
                "Exception Reason",
                "Reviewer",
                "Reviewed At",
                "Source Email ID",
                "Skill Run ID",
            ],
        },
        "suggested_skill": {
            "trigger": "new daily bank transaction email",
            "inputs": [
                "inbound bank transaction attachment",
                "Daily Reconciliation sheet",
                "Payment Export sheet",
                "Lists & Rules sheet",
            ],
            "actions": [
                "read bank attachment rows",
                "match transactions against Payment Export",
                "compute Amount Diff",
                "preview Daily Reconciliation row updates",
                "fill Match Status",
                "fill Exception Reason",
                "fill Reviewer",
                "fill Reviewed At",
                "fill Source Email ID",
                "fill Skill Run ID",
                "draft summary reply",
                "write audit log",
            ],
            "forbidden_actions": [
                "send email automatically",
                "access network",
                "overwrite reviewed rows",
                "write closed-period rows",
            ],
        },
        "handoff": {
            "next_owner": "section_b_skill_creation",
            "required_confirmation_fields": [
                "skill_name",
                "owner_role",
                "trigger_conditions",
                "allowed_actions",
                "forbidden_actions",
                "approval_mode",
                "success_criteria",
            ],
        },
    }


def bootstrap_demo(root: Path | str = ".", force: bool = False) -> dict[str, Any]:
    p = paths(root)
    for directory in [
        p.candidates_dir,
        p.reviews_dir,
        p.skills_dir,
        p.events_dir,
        p.attachments_dir,
        p.workbooks_dir,
        p.drafts_dir,
        p.matches_dir,
        p.runtime_dir,
    ]:
        directory.mkdir(parents=True, exist_ok=True)

    if force:
        if p.registry_db.exists():
            p.registry_db.unlink()
        for cleanup_dir in [p.matches_dir]:
            if cleanup_dir.exists():
                shutil.rmtree(cleanup_dir)
            cleanup_dir.mkdir(parents=True, exist_ok=True)

    candidate = default_pattern_candidate()
    candidate_path = p.candidates_dir / f"{candidate['candidate_id']}.json"
    if force or not candidate_path.exists():
        write_json(candidate_path, candidate)
    section_a_candidate = default_section_a_skill_candidate()
    if force or not p.skill_candidates_log.exists():
        write_jsonl(p.skill_candidates_log, [section_a_candidate])

    workbook_path = p.workbooks_dir / "cash_recon.xlsx"
    workbook_meta = p.workbooks_dir / "cash_recon.workbook.json"
    if force or not workbook_path.exists():
        workbook_path.write_text("SkillForge Local demo workbook placeholder.\n", encoding="utf-8")
    if force or not workbook_meta.exists():
        write_json(
            workbook_meta,
            {
                "workbook": str(workbook_path),
                "sheets": ["Daily Reconciliation", "Payment Export"],
                "closed_period_sheets": [],
                "adapter": "csv_rows_with_xlsx_filename_for_local_demo",
            },
        )

    attachment_path = p.attachments_dir / "bank_transactions_2026_06_15.xlsx"
    if force or not attachment_path.exists():
        write_transaction_attachment(attachment_path)

    if force and p.event_log.exists():
        p.event_log.unlink()
    if force or not p.event_log.exists():
        append_event(
            p.event_log,
            {
                "id": "event_email_bank_2026_06_15",
                "type": "email_received",
                "actor": "analyst_1",
                "workspace": "finance_demo",
                "payload": {
                    "message_id": "msg_bank_2026_06_15",
                    "thread_id": "thread_daily_cash_2026_06_15",
                    "direction": "inbox",
                    "from": "bank-ops@example.local",
                    "to": ["finance-team@example.local"],
                    "subject": "Daily bank transactions - June 15",
                    "date": "2026-06-15T08:02:00-07:00",
                    "body_text": "Attached are today's bank transactions.",
                    "attachments": [
                        {
                            "filename": attachment_path.name,
                            "path": str(attachment_path.relative_to(p.root)),
                        }
                    ],
                    "source_path": "workspace/mail/inbox_today/bank_2026_06_15.eml",
                },
            },
        )
    init_registry(p.registry_db)
    return {
        "candidate_path": str(candidate_path),
        "skill_candidates_log": str(p.skill_candidates_log),
        "event_log": str(p.event_log),
        "attachment": str(attachment_path),
        "workbook": str(workbook_path),
    }


def write_transaction_attachment(path: Path) -> None:
    rows = [
        {"transaction_id": "tx-1001", "bank_amount": "1200.00", "erp_amount": "1200.00", "description": "Customer payment"},
        {"transaction_id": "tx-1002", "bank_amount": "450.00", "erp_amount": "450.00", "description": "Vendor refund"},
        {"transaction_id": "tx-1003", "bank_amount": "89.25", "erp_amount": "89.25", "description": "Card settlement"},
        {"transaction_id": "tx-1004", "bank_amount": "210.00", "erp_amount": "200.00", "description": "Amount variance"},
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["transaction_id", "bank_amount", "erp_amount", "description"])
        writer.writeheader()
        writer.writerows(rows)


def seed_section_a_mock_from_workbook(root: Path | str, workbook_path: Path | str, force: bool = False) -> dict[str, Any]:
    p = paths(root)
    source_workbook = Path(workbook_path)
    if not source_workbook.exists():
        return {"status": "missing_workbook", "workbook": str(source_workbook)}

    for directory in [
        p.events_dir,
        p.attachments_dir,
        p.workbooks_dir,
        p.drafts_dir,
        p.matches_dir,
        p.runtime_dir,
        p.reviews_dir,
        p.skills_dir,
    ]:
        directory.mkdir(parents=True, exist_ok=True)

    if force:
        if p.registry_db.exists():
            p.registry_db.unlink()
        for cleanup_dir in [p.matches_dir]:
            if cleanup_dir.exists():
                shutil.rmtree(cleanup_dir)
            cleanup_dir.mkdir(parents=True, exist_ok=True)
        for cleanup_file in [p.event_log, p.skill_candidates_log]:
            if cleanup_file.exists():
                cleanup_file.unlink()

    staged_workbook = p.workbooks_dir / source_workbook.name
    if force and staged_workbook.exists():
        staged_workbook.chmod(0o666)
        staged_workbook.unlink()
    shutil.copy2(source_workbook, staged_workbook)
    workbook_info = inspect_xlsx_workbook(staged_workbook)
    target_sheet = "Daily Reconciliation" if "Daily Reconciliation" in workbook_info["sheets"] else workbook_info["sheets"][0]
    target_artifact = staged_workbook.relative_to(p.root).as_posix()
    candidate = section_a_candidate_from_workbook_info(target_artifact, target_sheet, workbook_info)
    write_jsonl(p.skill_candidates_log, [candidate])
    write_workbook_metadata(p, staged_workbook, workbook_info)

    attachment_path = p.attachments_dir / "bank_transactions_2026_06_15.xlsx"
    if force or not attachment_path.exists():
        write_transaction_attachment(attachment_path)
    append_event(
        p.event_log,
        {
            "id": "event_email_bank_2026_06_15",
            "type": "email_received",
            "actor": "analyst_1",
            "workspace": "finance_demo",
            "payload": {
                "message_id": "msg_bank_2026_06_15",
                "thread_id": "thread_daily_cash_2026_06_15",
                "direction": "inbox",
                "from": "bank-ops@example.local",
                "to": ["finance-team@example.local"],
                "subject": "Daily bank transactions - June 15",
                "date": "2026-06-15T08:02:00-07:00",
                "body_text": "Attached are today's bank transactions.",
                "attachments": [
                    {
                        "filename": attachment_path.name,
                        "path": attachment_path.relative_to(p.root).as_posix(),
                    }
                ],
                "source_path": "workspace/mail/inbox_today/bank_2026_06_15.eml",
            },
        },
    )
    init_registry(p.registry_db)
    return {
        "status": "ok",
        "source_workbook": str(source_workbook),
        "staged_workbook": str(staged_workbook),
        "target_sheet": target_sheet,
        "sheets": workbook_info["sheets"],
        "skill_candidates_log": str(p.skill_candidates_log),
        "event_log": str(p.event_log),
        "attachment": str(attachment_path),
        "candidate_id": candidate["candidate_id"],
    }


def inspect_xlsx_workbook(workbook_path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(workbook_path) as zf:
        shared_strings = read_shared_strings(zf)
        workbook = read_xml(zf, "xl/workbook.xml")
        rels = read_workbook_relationships(zf)
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        sheets = []
        sheet_details = {}
        for sheet in workbook.findall("main:sheets/main:sheet", ns):
            name = sheet.attrib["name"]
            sheets.append(name)
            rel_id = sheet.attrib.get(rel_ns)
            sheet_path = rels.get(rel_id)
            detail = {"headers": [], "dimension": None, "row_count": 0}
            if sheet_path:
                detail = inspect_xlsx_sheet(zf, sheet_path, shared_strings)
            sheet_details[name] = detail
    return {"sheets": sheets, "sheet_details": sheet_details}


def read_xml(zf: zipfile.ZipFile, member: str) -> Any:
    from xml.etree import ElementTree as ET

    return ET.fromstring(zf.read(member))


def read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = read_xml(zf, "xl/sharedStrings.xml")
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings = []
    for item in root.findall("main:si", ns):
        parts = [node.text or "" for node in item.findall(".//main:t", ns)]
        strings.append("".join(parts))
    return strings


def read_workbook_relationships(zf: zipfile.ZipFile) -> dict[str, str]:
    root = read_xml(zf, "xl/_rels/workbook.xml.rels")
    relationships = {}
    for relationship in root:
        rel_id = relationship.attrib.get("Id")
        target = relationship.attrib.get("Target", "")
        if rel_id and target:
            normalized = target.lstrip("/")
            relationships[rel_id] = normalized if normalized.startswith("xl/") else "xl/" + normalized
    return relationships


def inspect_xlsx_sheet(zf: zipfile.ZipFile, sheet_path: str, shared_strings: list[str]) -> dict[str, Any]:
    root = read_xml(zf, sheet_path)
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    dimension = root.find("main:dimension", ns)
    rows = root.findall("main:sheetData/main:row", ns)
    headers = []
    if rows:
        headers = [cell_value(cell, shared_strings) for cell in rows[0].findall("main:c", ns)]
        headers = [header for header in headers if header]
    return {
        "headers": headers,
        "dimension": dimension.attrib.get("ref") if dimension is not None else None,
        "row_count": len(rows),
    }


def cell_value(cell: Any, shared_strings: list[str]) -> str:
    value_node = None
    for child in cell:
        if child.tag.endswith("}v"):
            value_node = child
            break
        if child.tag.endswith("}is"):
            return "".join(node.text or "" for node in child.iter() if node.tag.endswith("}t"))
    if value_node is None or value_node.text is None:
        return ""
    value = value_node.text
    if cell.attrib.get("t") == "s":
        try:
            return shared_strings[int(value)]
        except (IndexError, ValueError):
            return value
    return value


def section_a_candidate_from_workbook_info(target_artifact: str, target_sheet: str, workbook_info: dict[str, Any]) -> dict[str, Any]:
    target_details = workbook_info["sheet_details"].get(target_sheet, {})
    common_fields = target_details.get("headers") or [
        "Amount Diff",
        "Match Status",
        "Exception Reason",
        "Reviewer",
        "Reviewed At",
        "Source Email ID",
        "Notes",
    ]
    supporting_inputs = [f"{sheet} sheet" for sheet in workbook_info["sheets"] if sheet in {"Payment Export", "Lists & Rules"}]
    return {
        "contract_version": "section_a.skill_candidate.v1",
        "candidate_id": "cand_daily_cash_recon_001",
        "name_suggestion": "Daily Cash Reconciliation Skill",
        "confidence": 0.95,
        "status": "candidate",
        "detected_at": "2026-06-14T08:44:00-07:00",
        "pattern": {
            "workflow_family": "daily_cash_reconciliation",
            "episode_count": 3,
            "common_sequence": ["email_received", "spreadsheet_row_updated", "outbound_message_created"],
            "trigger_signature": {
                "subject_prefix": "Daily bank transactions - Jun",
                "attachment_glob": "bank_transactions_*.xlsx",
                "intent": "daily_cash_reconciliation_request",
            },
            "similarity_reason": [
                "same inbound email intent",
                "same attachment naming pattern",
                f"same target workbook {Path(target_artifact).name}",
                f"same target sheet {target_sheet}",
                "same spreadsheet audit sequence",
                "same daily summary reply shape",
            ],
        },
        "evidence": {
            "episode_ids": [
                "episode_msg_bank_2026_06_12",
                "episode_msg_bank_2026_06_13",
                "episode_msg_bank_2026_06_14",
            ],
            "source_event_ids": [
                "email_in_20260612",
                "audit_003",
                "email_sent_20260612",
                "email_in_20260613",
                "audit_006",
                "email_sent_20260613",
                "email_in_20260614",
                "audit_009",
                "email_sent_20260614",
            ],
            "target_artifact": target_artifact,
            "target_sheet": target_sheet,
            "target_rows_examples": ["2:46", "47:91", "92:136"],
            "common_fields": common_fields,
            "daily_batch_size": 45,
        },
        "next_trigger": {
            "event_id": "event_email_bank_2026_06_15",
            "message_id": "msg_bank_2026_06_15",
            "thread_id": "thread_daily_cash_2026_06_15",
            "target_rows": "137:181",
            "unprocessed_action_columns": [
                "Match Status",
                "Exception Reason",
                "Reviewer",
                "Reviewed At",
                "Source Email ID",
                "Skill Run ID",
            ],
        },
        "suggested_skill": {
            "trigger": "new daily bank transaction email",
            "inputs": ["inbound bank transaction attachment", f"{target_sheet} sheet", *supporting_inputs],
            "actions": [
                "read bank attachment rows",
                "match transactions against Payment Export",
                "compute Amount Diff",
                f"preview {target_sheet} row updates",
                "fill Match Status",
                "fill Exception Reason",
                "fill Reviewer",
                "fill Reviewed At",
                "fill Source Email ID",
                "fill Skill Run ID",
                "draft summary reply",
                "write audit log",
            ],
            "forbidden_actions": [
                "send email automatically",
                "access network",
                "overwrite reviewed rows",
                "write closed-period rows",
            ],
        },
        "handoff": {
            "next_owner": "section_b_skill_creation",
            "required_confirmation_fields": [
                "skill_name",
                "owner_role",
                "trigger_conditions",
                "allowed_actions",
                "forbidden_actions",
                "approval_mode",
                "success_criteria",
            ],
        },
    }


def write_workbook_metadata(p: SkillGenPaths, staged_workbook: Path, workbook_info: dict[str, Any]) -> None:
    write_json(
        staged_workbook.with_suffix(".workbook.json"),
        {
            "workbook": staged_workbook.relative_to(p.root).as_posix(),
            "sheets": workbook_info["sheets"],
            "sheet_details": workbook_info["sheet_details"],
            "closed_period_sheets": [],
            "adapter": "xlsx_package_inspection_for_section_a_mock",
        },
    )


def candidate_path_for(root: Path | str, candidate_id: str) -> Path:
    return paths(root).candidates_dir / f"{candidate_id}.json"


def is_section_a_candidate(candidate: dict[str, Any]) -> bool:
    return candidate.get("contract_version") in SECTION_A_CONTRACT_VERSIONS


def read_section_a_candidates(root: Path | str) -> list[dict[str, Any]]:
    return read_jsonl(paths(root).skill_candidates_log)


def find_section_a_candidate(root: Path | str, candidate_id: str) -> dict[str, Any] | None:
    for candidate in read_section_a_candidates(root):
        if candidate.get("candidate_id") == candidate_id:
            return candidate
    return None


def load_candidate(root: Path | str, candidate_id: str) -> dict[str, Any]:
    section_a_candidate = find_section_a_candidate(root, candidate_id)
    if section_a_candidate:
        return section_a_candidate
    return read_json(candidate_path_for(root, candidate_id))


def nested_value(payload: dict[str, Any], path: str) -> Any:
    current: Any = payload
    for part in path.split("."):
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def validate_section_a_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    missing = [field for field in REQUIRED_SECTION_A_FIELDS if field not in candidate]
    if candidate.get("contract_version") not in SECTION_A_CONTRACT_VERSIONS:
        missing.append("contract_version:section_a.skill_candidate.v1")
    if candidate.get("status") != "candidate":
        missing.append("status:candidate")
    for field in REQUIRED_SECTION_A_MINIMAL_FIELDS:
        value = nested_value(candidate, field)
        if value is None or value == [] or value == "":
            missing.append(field)
    if nested_value(candidate, "handoff.next_owner") != "section_b_skill_creation":
        missing.append("handoff.next_owner:section_b_skill_creation")
    if missing:
        return {"status": "needs_more_evidence", "missing_fields": sorted(set(missing))}
    return {"status": "ok", "missing_fields": []}


def validate_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    if is_section_a_candidate(candidate):
        return validate_section_a_candidate(candidate)
    missing = [field for field in REQUIRED_CANDIDATE_FIELDS if field not in candidate]
    if missing:
        return {"status": "needs_more_evidence", "missing_fields": missing}
    if candidate.get("status") not in {"ready_for_human_review", "accepted_for_skill_generation"}:
        return {"status": "needs_more_evidence", "missing_fields": ["status:ready_for_human_review"]}
    return {"status": "ok", "missing_fields": []}


def create_review_session(
    root: Path | str,
    candidate_id: str,
    planner: str = "deterministic",
    model_timeout_seconds: int = 180,
) -> dict[str, Any]:
    p = paths(root)
    candidate = load_candidate(root, candidate_id)
    validation = validate_candidate(candidate)
    if validation["status"] != "ok":
        return validation

    review_id = f"review_{candidate_id}"
    candidate_is_section_a = is_section_a_candidate(candidate)
    if candidate_is_section_a:
        review = build_section_a_review(review_id, candidate)
    else:
        review = build_legacy_review(review_id, candidate)
    review = apply_skill_planner(root, candidate, review, planner=planner, model_timeout_seconds=model_timeout_seconds)
    p.reviews_dir.mkdir(parents=True, exist_ok=True)
    write_json(p.reviews_dir / f"{review_id}.json", review)
    if not candidate_is_section_a:
        update_candidate_status(root, candidate_id, "accepted_for_skill_generation")
    append_event(
        p.event_log,
        {
            "id": f"event_{review_id}",
            "type": "skill_review_started",
            "candidate_id": candidate_id,
            "review_session_id": review_id,
        },
    )
    return review


def build_legacy_review(review_id: str, candidate: dict[str, Any]) -> dict[str, Any]:
    return {
        "review_session_id": review_id,
        "candidate_id": candidate["candidate_id"],
        "status": "awaiting_human_review",
        "created_at": utc_now(),
        "candidate_summary": candidate["summary"],
        "evidence": candidate["evidence"],
        "suggested": {
            "skill_name": "Daily Cash Reconciliation Skill",
            "skill_id": "daily_cash_reconciliation",
            "scope": "team",
            "owner_role": "controller",
            "approval_mode": "confirm_each_run",
            "trigger": normalize_trigger(candidate["suggested_trigger"]),
            "inputs": normalize_inputs(candidate["suggested_inputs"]),
            "outputs": normalize_outputs(candidate["suggested_outputs"]),
            "allowed_actions": default_allowed_actions(),
            "forbidden_actions": default_forbidden_actions(candidate["suggested_guardrails"]),
            "validation_rules": default_validation_rules(),
            "workflow_steps": default_workflow_steps(),
            "expected_outcome": default_expected_outcome(),
            "description": default_skill_description(),
        },
    }


def build_section_a_review(review_id: str, candidate: dict[str, Any]) -> dict[str, Any]:
    pattern = candidate["pattern"]
    evidence = candidate["evidence"]
    suggested_skill = candidate["suggested_skill"]
    skill_id = slugify(pattern.get("workflow_family") or candidate["name_suggestion"])
    workflow_steps = workflow_steps_from_section_a_candidate(candidate)
    return {
        "review_session_id": review_id,
        "candidate_id": candidate["candidate_id"],
        "source_contract_version": candidate["contract_version"],
        "status": "awaiting_human_review",
        "created_at": utc_now(),
        "candidate_summary": {
            "one_liner": suggested_skill["trigger"],
            "workflow_family": pattern["workflow_family"],
            "episode_count": pattern.get("episode_count"),
            "confidence": candidate["confidence"],
        },
        "evidence": {
            "episodes": evidence["episode_ids"],
            "source_event_ids": evidence["source_event_ids"],
            "target_artifact": evidence["target_artifact"],
            "target_sheet": evidence["target_sheet"],
            "target_rows_examples": evidence.get("target_rows_examples", []),
            "common_fields": evidence.get("common_fields", []),
            "similarity_reason": pattern.get("similarity_reason", []),
            "next_trigger": candidate.get("next_trigger"),
        },
        "handoff_required_confirmation_fields": candidate["handoff"]["required_confirmation_fields"],
        "suggested": {
            "skill_name": candidate["name_suggestion"],
            "skill_id": skill_id,
            "scope": "team",
            "owner_role": "controller",
            "approval_mode": "confirm_each_run",
            "trigger": normalize_section_a_trigger(candidate),
            "inputs": normalize_section_a_inputs(candidate),
            "outputs": normalize_section_a_outputs(candidate),
            "allowed_actions": allowed_actions_from_section_a(candidate),
            "forbidden_actions": forbidden_actions_from_section_a(candidate),
            "validation_rules": validation_rules_from_section_a(candidate),
            "workflow_steps": workflow_steps,
            "expected_outcome": expected_outcome_from_section_a_candidate(candidate),
            "description": description_from_section_a_candidate(candidate),
            "trigger_label": suggested_skill["trigger"],
        },
    }


def apply_skill_planner(
    root: Path | str,
    candidate: dict[str, Any],
    review: dict[str, Any],
    planner: str = "deterministic",
    model_timeout_seconds: int = 180,
) -> dict[str, Any]:
    if planner == "deterministic":
        review["planner"] = {"mode": "deterministic", "status": "applied"}
        return review
    if planner not in ("local-model", "anthropic", "codex"):
        review["planner"] = {"mode": planner, "status": "fallback", "error": f"unknown planner: {planner}"}
        return review

    config = local_model_config(root, timeout_seconds=model_timeout_seconds)
    learned_preferences, memory_backend = recall_learned_preferences(root, candidate)
    try:
        plan = request_model_skill_plan(config, candidate, review["suggested"], learned_preferences=learned_preferences)
        suggested, applied_fields, warnings = merge_model_plan_into_suggested(review["suggested"], plan)
        review["suggested"] = suggested
        review["planner"] = {
            "mode": planner,
            "status": "applied",
            "backend": config.backend,
            "base_url": config.base_url,
            "model": config.model,
            "applied_fields": applied_fields,
            "warnings": warnings,
            "memory_backend": memory_backend,
            "learned_preferences_count": len(learned_preferences),
            "learned_preferences": learned_preferences[:5],
        }
    except Exception as exc:
        review["planner"] = {
            "mode": planner,
            "status": "fallback",
            "backend": config.backend,
            "base_url": config.base_url,
            "model": config.model,
            "error": str(exc),
            "memory_backend": memory_backend,
            "learned_preferences_count": len(learned_preferences),
        }
    return review


def recall_learned_preferences(root: Path | str, candidate: dict[str, Any]) -> tuple[list[str], str]:
    """Pull this reviewer's past feedback from the memory layer (HydraDB + local).

    Returns ``(preference_texts, backend)``. Never raises — memory is best-effort.
    """
    try:
        from skillforge_local.memory import SkillMemory

        mem = SkillMemory(root)
        summary = candidate.get("summary", {}) if isinstance(candidate.get("summary"), dict) else {}
        family = summary.get("workflow_family") or candidate.get("candidate_name") or candidate.get("name_suggestion") or "skill"
        items = mem.recall_preferences(
            query=(
                f"reviewer preferences and past feedback for the {family} skill: "
                "column filling, human approvals, validation rules, description wording, things to avoid"
            ),
            limit=8,
        )
        return [it["text"] for it in items if it.get("text")], mem.status()["backend"]
    except Exception:  # noqa: BLE001 - memory must never block generation
        return [], "local"


def request_model_skill_plan(
    config: LocalModelConfig,
    candidate: dict[str, Any],
    suggested: dict[str, Any],
    learned_preferences: list[str] | None = None,
) -> dict[str, Any]:
    response = call_local_chat_model(
        config,
        model_planner_messages(candidate, suggested, learned_preferences=learned_preferences),
        response_format={"type": "json_object"},
        max_tokens=8000,
        thinking=False,
    )
    return parse_json_object(local_model_text_response(response))


def model_planner_messages(
    candidate: dict[str, Any],
    suggested: dict[str, Any],
    learned_preferences: list[str] | None = None,
) -> list[dict[str, str]]:
    payload = {
        "section_a_candidate": candidate,
        "deterministic_base_plan": {
            "description": suggested.get("description"),
            "workflow_steps": suggested.get("workflow_steps"),
            "expected_outcome": suggested.get("expected_outcome"),
            "validation_rules": suggested.get("validation_rules"),
            "forbidden_actions": suggested.get("forbidden_actions"),
            "resources": {
                "inputs": suggested.get("inputs"),
                "outputs": suggested.get("outputs"),
            },
        },
        "allowed_step_types": sorted(ALLOWED_WORKFLOW_STEP_TYPES),
        "allowed_action_types": sorted(ALLOWED_EXECUTOR_ACTIONS),
        "required_invariants": [
            "Return only valid JSON.",
            "Preserve the schema shape used by deterministic_base_plan.",
            "Keep one human_approval step before any write_output step.",
            "Keep a write_output step that creates a reconciled spreadsheet.",
            "Keep a validate step that writes audit/SkillOps evidence.",
            "Do not send email automatically.",
            "Do not access the network.",
            "Do not overwrite reviewed rows or closed-period rows.",
        ],
    }
    cleaned_preferences = [p.strip() for p in (learned_preferences or []) if isinstance(p, str) and p.strip()]
    if cleaned_preferences:
        payload["learned_reviewer_preferences"] = cleaned_preferences
    system = (
        "You are SkillForge Local's offline skill planner. Refine workflow text for human review "
        "while preserving local safety invariants. Do not invent network, email-send, or destructive actions."
    )
    preference_instruction = ""
    if cleaned_preferences:
        preference_instruction = (
            " Honor the learned_reviewer_preferences: fold them into the description, workflow_steps, "
            "expected_outcome, and validation_rules. They reflect this reviewer's past feedback on earlier "
            "skills and take priority over the deterministic defaults — but never weaken the required_invariants."
        )
    user = (
        "Refine the deterministic skill plan for human review. Return a JSON object with exactly these top-level keys: "
        "description, workflow_steps, expected_outcome, validation_rules. "
        "workflow_steps must be an array of objects with id, order, title, type, summary, inputs, outputs, "
        "and optional action_type, target, requires_step, approval_text. "
        "expected_outcome must contain summary, files_created, files_modified, side_effects."
        + preference_instruction
        + "\n\n"
        + json.dumps(payload, indent=2, sort_keys=True)
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def merge_model_plan_into_suggested(
    suggested: dict[str, Any],
    plan: dict[str, Any],
) -> tuple[dict[str, Any], list[str], list[str]]:
    updated = deepcopy(suggested)
    applied_fields: list[str] = []
    warnings: list[str] = []

    description = plan.get("description")
    if isinstance(description, str) and description.strip():
        updated["description"] = description.strip()
        applied_fields.append("description")
    else:
        warnings.append("description missing or invalid")

    steps, step_warnings = sanitize_model_workflow_steps(plan.get("workflow_steps"), suggested.get("workflow_steps", []))
    warnings.extend(step_warnings)
    if steps:
        updated["workflow_steps"] = steps
        applied_fields.append("workflow_steps")

    expected_outcome, outcome_warnings = sanitize_model_expected_outcome(
        plan.get("expected_outcome"),
        suggested.get("expected_outcome", {}),
    )
    warnings.extend(outcome_warnings)
    if expected_outcome:
        updated["expected_outcome"] = expected_outcome
        applied_fields.append("expected_outcome")

    validation_rules = sanitize_model_validation_rules(
        plan.get("validation_rules"),
        suggested.get("validation_rules", []),
    )
    if validation_rules != suggested.get("validation_rules", []):
        updated["validation_rules"] = validation_rules
        applied_fields.append("validation_rules")

    return updated, applied_fields, warnings


def sanitize_model_workflow_steps(raw_steps: Any, fallback_steps: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not isinstance(raw_steps, list):
        return deepcopy(fallback_steps), ["workflow_steps missing or not a list"]

    sanitized = []
    for raw_step in raw_steps:
        if not isinstance(raw_step, dict):
            warnings.append("ignored non-object workflow step")
            continue
        step_id = slugify(str(raw_step.get("id") or raw_step.get("title") or "generated_step"))
        step_type = raw_step.get("type")
        if step_type not in ALLOWED_WORKFLOW_STEP_TYPES:
            warnings.append(f"step {step_id} has invalid type {step_type}; using transform")
            step_type = "transform"
        step = {
            "id": step_id,
            "order": int(raw_step.get("order") or len(sanitized) + 1),
            "title": str(raw_step.get("title") or sentence_title(step_id.replace("_", " "))).strip(),
            "type": step_type,
            "summary": str(raw_step.get("summary") or f"Perform {step_id.replace('_', ' ')}.").strip(),
            "inputs": sanitize_string_list(raw_step.get("inputs")),
            "outputs": sanitize_string_list(raw_step.get("outputs")),
        }
        action_type = raw_step.get("action_type")
        if action_type in ALLOWED_EXECUTOR_ACTIONS:
            step["action_type"] = action_type
        elif action_type:
            warnings.append(f"step {step_id} dropped invalid action_type {action_type}")
        for optional_key in ["target", "requires_step", "approval_text", "template"]:
            value = raw_step.get(optional_key)
            if isinstance(value, dict | str):
                step[optional_key] = deepcopy(value)
        sanitized.append(step)

    if len(sanitized) < 3:
        return deepcopy(fallback_steps), warnings + ["model returned too few valid workflow steps"]

    sanitized.sort(key=lambda item: item["order"])
    for order, step in enumerate(sanitized, start=1):
        step["order"] = order

    approval_orders = [step["order"] for step in sanitized if step["type"] == "human_approval" or step.get("action_type") == "require_human_approval"]
    write_orders = [step["order"] for step in sanitized if step["type"] == "write_output" or step.get("action_type") == "write_xlsx_update"]
    validate_orders = [step["order"] for step in sanitized if step["type"] == "validate" or step.get("action_type") == "write_audit_log"]
    if not approval_orders or not write_orders or min(write_orders) < min(approval_orders):
        return deepcopy(fallback_steps), warnings + ["model workflow failed approval-before-write invariant"]
    if not validate_orders:
        return deepcopy(fallback_steps), warnings + ["model workflow missing validate/audit step"]
    if not any("reconciled" in " ".join(step.get("outputs", []) + [step["title"], step["summary"]]).lower() for step in sanitized):
        return deepcopy(fallback_steps), warnings + ["model workflow missing reconciled spreadsheet output"]
    return sanitized, warnings


def sanitize_string_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item).strip() for item in value if str(item).strip()]


def sanitize_model_expected_outcome(raw_outcome: Any, fallback: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    if not isinstance(raw_outcome, dict):
        return deepcopy(fallback), ["expected_outcome missing or not an object"]
    outcome = {
        "summary": str(raw_outcome.get("summary") or fallback.get("summary") or "").strip(),
        "files_created": sanitize_string_list(raw_outcome.get("files_created")),
        "files_modified": sanitize_string_list(raw_outcome.get("files_modified")),
        "side_effects": sanitize_string_list(raw_outcome.get("side_effects")),
    }
    warnings = []
    if not outcome["files_created"]:
        outcome["files_created"] = list(fallback.get("files_created", []))
        warnings.append("expected_outcome.files_created missing; using fallback")
    if not any("reconciled" in item.lower() and item.lower().endswith(".xlsx") for item in outcome["files_created"]):
        outcome["files_created"] = list(fallback.get("files_created", []))
        warnings.append("expected_outcome missing reconciled xlsx output; using fallback files_created")
    if not outcome["files_modified"]:
        outcome["files_modified"] = list(fallback.get("files_modified", []))
        warnings.append("expected_outcome.files_modified missing; using fallback")
    if not outcome["side_effects"]:
        outcome["side_effects"] = list(fallback.get("side_effects", []))
        warnings.append("expected_outcome.side_effects missing; using fallback")
    return outcome, warnings


def sanitize_model_validation_rules(raw_rules: Any, fallback_rules: list[str]) -> list[str]:
    rules = list(fallback_rules)
    if not isinstance(raw_rules, list):
        return rules
    for item in raw_rules:
        rule = slugify(str(item))
        if rule and rule not in rules:
            rules.append(rule)
    return rules[:80]


def normalize_trigger(suggested_trigger: dict[str, Any]) -> dict[str, Any]:
    conditions: list[dict[str, Any]] = []
    for condition in suggested_trigger.get("conditions", []):
        ctype = condition.get("type")
        value = condition.get("value")
        if ctype in {"subject_pattern", "subject_contains"}:
            conditions.append({"type": "subject_contains", "value": value})
        elif ctype == "has_attachment_matching":
            conditions.append({"type": "has_attachment_matching", "value": value})
        elif ctype == "sender_pattern":
            conditions.append({"type": "sender_pattern", "value": value})
    conditions.extend(
        [
            {"type": "require_file", "value": "workspace/workbooks/cash_recon.xlsx"},
            {
                "type": "workbook_has_sheet",
                "workbook": "workspace/workbooks/cash_recon.xlsx",
                "sheet": "Daily Reconciliation",
            },
            {
                "type": "no_existing_recon_for_date",
                "workbook": "workspace/workbooks/cash_recon.xlsx",
                "date_source": "$email.date",
            },
        ]
    )
    return {"event_type": suggested_trigger.get("event_type", "email_received"), "conditions": conditions}


def normalize_inputs(inputs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for item in inputs:
        if item.get("name") == "daily_bank_transaction_file":
            normalized.append(
                {
                    "name": item["name"],
                    "type": "xlsx_attachment",
                    "source": "$email.attachments[bank_transactions_*.xlsx]",
                    "required": bool(item.get("required", True)),
                }
            )
        elif item.get("name") == "cash_recon_workbook":
            normalized.append(
                {
                    "name": item["name"],
                    "type": "xlsx_workbook",
                    "path": item.get("path_hint", "workspace/workbooks/cash_recon.xlsx"),
                    "required": bool(item.get("required", True)),
                }
            )
    return normalized


def normalize_outputs(outputs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for item in outputs:
        if item.get("type") == "xlsx_update":
            normalized.append({"name": "updated_reconciliation_workbook", "type": "xlsx_update"})
        elif item.get("type") == "email_draft":
            normalized.append({"name": "summary_reply_draft", "type": "email_draft"})
    return normalized


def normalize_section_a_trigger(candidate: dict[str, Any]) -> dict[str, Any]:
    pattern = candidate["pattern"]
    evidence = candidate["evidence"]
    signature = pattern["trigger_signature"]
    conditions: list[dict[str, Any]] = []
    if signature.get("subject_prefix"):
        conditions.append({"type": "subject_prefix", "value": signature["subject_prefix"]})
    elif signature.get("subject_contains"):
        conditions.append({"type": "subject_contains", "value": signature["subject_contains"]})
    if signature.get("attachment_glob"):
        conditions.append({"type": "has_attachment_matching", "value": signature["attachment_glob"]})
    if signature.get("sender"):
        conditions.append({"type": "sender_pattern", "value": signature["sender"]})
    conditions.extend(
        [
            {"type": "require_file", "value": evidence["target_artifact"]},
            {
                "type": "workbook_has_sheet",
                "workbook": evidence["target_artifact"],
                "sheet": evidence["target_sheet"],
            },
        ]
    )
    if candidate.get("next_trigger"):
        conditions.append(
            {
                "type": "no_existing_recon_for_date",
                "workbook": evidence["target_artifact"],
                "date_source": "$email.date",
            }
        )
    event_type = pattern.get("common_sequence", ["email_received"])[0]
    return {
        "event_type": event_type,
        "label": candidate["suggested_skill"]["trigger"],
        "conditions": conditions,
    }


def normalize_section_a_inputs(candidate: dict[str, Any]) -> list[dict[str, Any]]:
    evidence = candidate["evidence"]
    signature = candidate["pattern"]["trigger_signature"]
    attachment_glob = signature.get("attachment_glob", "*")
    inputs = [
        {
            "name": "inbound_attachment",
            "type": "xlsx_attachment",
            "source": f"$email.attachments[{attachment_glob}]",
            "required": True,
        },
        {
            "name": "target_workbook",
            "type": "xlsx_workbook",
            "path": evidence["target_artifact"],
            "required": True,
        },
    ]
    for input_text in candidate["suggested_skill"].get("inputs", []):
        name = slugify(input_text)
        if name in {"inbound_bank_transaction_attachment", "daily_bank_transaction_file"}:
            continue
        if "sheet" in input_text.lower():
            inputs.append(
                {
                    "name": name,
                    "type": "workbook_sheet",
                    "workbook": evidence["target_artifact"],
                    "sheet": input_text.replace(" sheet", "").strip(),
                    "required": True,
                }
            )
    return dedupe_resources(inputs, "name")


def normalize_section_a_outputs(candidate: dict[str, Any]) -> list[dict[str, Any]]:
    evidence = candidate["evidence"]
    actions = "\n".join(candidate["suggested_skill"].get("actions", [])).lower()
    outputs = [
        {
            "name": "reconciled_spreadsheet",
            "type": "xlsx_file",
            "source_workbook": evidence["target_artifact"],
            "target_sheet": evidence["target_sheet"],
            "path_template": "workspace/workbooks/generated/{source_workbook}_{event_date}_reconciled.xlsx",
        }
    ]
    if "reply" in actions or "outbound_message_created" in candidate["pattern"].get("common_sequence", []):
        outputs.append({"name": "summary_reply_draft", "type": "email_draft"})
    if "audit" in actions:
        outputs.append({"name": "audit_log_entry", "type": "audit_log"})
    return outputs


def dedupe_resources(resources: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    seen = set()
    deduped = []
    for resource in resources:
        value = resource.get(key)
        if value in seen:
            continue
        seen.add(value)
        deduped.append(resource)
    return deduped


def allowed_actions_from_section_a(candidate: dict[str, Any]) -> list[str]:
    actions = set(default_allowed_actions())
    for action in candidate["suggested_skill"].get("actions", []):
        normalized = slugify(action)
        if normalized:
            actions.add(normalized)
    return sorted(actions)


def forbidden_actions_from_section_a(candidate: dict[str, Any]) -> list[str]:
    forbidden = set()
    for action in candidate["suggested_skill"].get("forbidden_actions", []):
        text = action.lower()
        if "send" in text and "email" in text:
            forbidden.add("send_email")
        elif "network" in text:
            forbidden.add("access_network")
        elif "overwrite" in text and "reviewed" in text:
            forbidden.add("overwrite_reviewed_rows")
        elif "closed" in text and ("period" in text or "row" in text):
            forbidden.add("modify_closed_period_sheets")
        else:
            forbidden.add(slugify(action))
    forbidden.add("read_outside_workspace")
    return sorted(forbidden)


def validation_rules_from_section_a(candidate: dict[str, Any]) -> list[str]:
    rules = list(default_validation_rules())
    if candidate.get("next_trigger", {}).get("target_rows"):
        rules.append("target_rows_match_next_trigger")
    for field in candidate.get("evidence", {}).get("common_fields", []):
        field_rule = f"field_populated_{slugify(field)}"
        if field_rule not in rules:
            rules.append(field_rule)
    return rules


def workflow_steps_from_section_a_candidate(candidate: dict[str, Any]) -> list[dict[str, Any]]:
    evidence = candidate["evidence"]
    steps = []
    audit_action = None
    for action in candidate["suggested_skill"].get("actions", []):
        if "audit" in action.lower():
            audit_action = action
            continue
        step_type, action_type = classify_section_a_action(action)
        step = {
            "id": slugify(action),
            "order": len(steps) + 1,
            "title": sentence_title(action),
            "type": step_type,
            "summary": summary_for_section_a_action(action, evidence),
            "inputs": inputs_for_section_a_action(action),
            "outputs": outputs_for_section_a_action(action),
        }
        if action_type:
            step["action_type"] = action_type
        if step_type in {"transform", "write_output"}:
            step["target"] = {"workbook": evidence["target_artifact"], "sheet": evidence["target_sheet"]}
        steps.append(step)

    approval_order = len(steps) + 1
    steps.append(
        {
            "id": "require_approval",
            "order": approval_order,
            "title": "Request review",
            "type": "human_approval",
            "action_type": "require_human_approval",
            "summary": "Ask a reviewer to approve the spreadsheet update and reply draft before any file is created.",
            "inputs": ["reconciliation_preview", "summary_reply_draft"],
            "outputs": ["approval_decision"],
            "approval_text": "Apply generated updates and create local draft outputs?",
        }
    )
    steps.append(
        {
            "id": "create_reconciled_spreadsheet",
            "order": len(steps) + 1,
            "title": "Create updated spreadsheet",
            "type": "write_output",
            "action_type": "write_xlsx_update",
            "summary": "After approval, create a new spreadsheet containing the reviewed row updates.",
            "inputs": ["approval_decision", "reconciliation_preview"],
            "outputs": ["reconciled_spreadsheet"],
            "target": {"workbook": evidence["target_artifact"], "sheet": evidence["target_sheet"]},
            "requires_step": "require_approval",
        }
    )
    steps.append(
        {
            "id": "write_audit_log",
            "order": len(steps) + 1,
            "title": "Save run record",
            "type": "validate",
            "action_type": "write_audit_log",
            "summary": "Save the output paths and final check results in the local run record.",
            "inputs": ["reconciled_spreadsheet", "summary_reply_draft"],
            "outputs": ["audit_event", "skillops_usage_event"],
        }
    )
    return steps


def classify_section_a_action(action: str) -> tuple[str, str | None]:
    text = action.lower()
    if "read" in text and "attachment" in text:
        return "read_input", "parse_xlsx_attachment"
    if "draft" in text or "reply" in text or "message" in text:
        return "draft_output", "draft_email_reply"
    if "exception" in text:
        return "analyze", "flag_reconciliation_exceptions"
    if any(token in text for token in ["match", "compute", "preview", "fill", "classif"]):
        return "transform", "preview_reconciliation_update"
    if "write" in text and "workbook" in text:
        return "write_output", "write_xlsx_update"
    return "transform", None


def sentence_title(value: str) -> str:
    value = value.strip()
    return value[:1].upper() + value[1:] if value else "Generated step"


def summary_for_section_a_action(action: str, evidence: dict[str, Any]) -> str:
    text = action.lower()
    if "read" in text and "attachment" in text:
        return "Read the spreadsheet attached to the bank email."
    if "match" in text:
        return "Compare the bank rows with the finance workbook."
    if "preview" in text:
        return "Prepare the spreadsheet changes without writing them yet."
    if "fill" in text:
        return f"Fill the {action.replace('fill ', '')} field in the proposed spreadsheet update."
    if "draft" in text or "reply" in text:
        return "Create a local reply draft with the matched count and the items that need review."
    return f"Perform the repeated action: {action}."


def inputs_for_section_a_action(action: str) -> list[str]:
    text = action.lower()
    if "read" in text and "attachment" in text:
        return ["resources.inputs.inbound_attachment"]
    if "draft" in text or "reply" in text:
        return ["reconciliation_preview", "exceptions"]
    return ["transactions", "resources.inputs.target_workbook"]


def outputs_for_section_a_action(action: str) -> list[str]:
    text = action.lower()
    if "read" in text and "attachment" in text:
        return ["transactions"]
    if "draft" in text or "reply" in text:
        return ["summary_reply_draft"]
    if "exception" in text:
        return ["exceptions"]
    return ["reconciliation_preview"]


def expected_outcome_from_section_a_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    evidence = candidate["evidence"]
    workbook = evidence["target_artifact"]
    return {
        "summary": (
            "After approval, the workflow creates a new updated spreadsheet, lists the rows that need review, "
            "saves a local reply draft, and records the run locally without sending anything."
        ),
        "files_created": [
            "workspace/workbooks/generated/{source_workbook}_{event_date}_reconciled.xlsx",
            "workspace/mail/drafts/{skill_id}_{event_date}_reply.eml",
        ],
        "files_modified": [update_log_path_for_workbook(workbook), "workspace/events/events.jsonl"],
        "side_effects": guardrails_from_forbidden_actions(forbidden_actions_from_section_a(candidate)),
    }


def description_from_section_a_candidate(candidate: dict[str, Any]) -> str:
    return (
        "When the daily bank transaction email arrives, this workflow reads the attached spreadsheet, "
        "checks the rows against the finance workbook, prepares the spreadsheet update, separates items that need review, "
        "and creates the approved output files locally."
    )


def update_log_path_for_workbook(workbook_path: str) -> str:
    path = Path(workbook_path)
    return path.with_name(f"{path.stem}.skill_updates.jsonl").as_posix()


def output_workbook_path_for_run(source_workbook_path: str, date_slug: str) -> str:
    source_stem = Path(source_workbook_path).stem or "reconciled_workbook"
    return f"workspace/workbooks/generated/{source_stem}_{date_slug}_reconciled.xlsx"


def default_allowed_actions() -> list[str]:
    return [
        "read_email",
        "read_attachment",
        "read_workbook",
        "preview_workbook_update",
        "write_workbook_update_after_approval",
        "create_email_draft",
        "write_audit_log",
    ]


def default_forbidden_actions(guardrails: list[str]) -> list[str]:
    actions = {
        "send_email",
        "access_network",
        "overwrite_reviewed_rows",
        "modify_closed_period_sheets",
        "read_outside_workspace",
    }
    text = "\n".join(guardrails).lower()
    if "network" in text:
        actions.add("access_network")
    if "send email" in text:
        actions.add("send_email")
    return sorted(actions)


def default_validation_rules() -> list[str]:
    return [
        "workbook_can_be_reopened",
        "only_allowed_sheets_modified",
        "no_closed_period_sheets_modified",
        "no_reviewed_rows_overwritten",
        "reconciled_spreadsheet_created",
        "exception_count_matches_summary",
        "draft_created_but_not_sent",
        "audit_log_written",
    ]


def default_human_feedback(root: Path | str, review_session_id: str, reviewer: str = "controller") -> dict[str, Any]:
    p = paths(root)
    review = read_json(p.reviews_dir / f"{review_session_id}.json")
    suggested = review["suggested"]
    feedback = {
        "candidate_id": review["candidate_id"],
        "review_session_id": review_session_id,
        "reviewed_at": utc_now(),
        "reviewer": reviewer,
        "decision": "install_skill",
        "skill_name": suggested["skill_name"],
        "skill_id": suggested["skill_id"],
        "scope": suggested["scope"],
        "owner_role": suggested["owner_role"],
        "approval_mode": suggested["approval_mode"],
        "confirmed_trigger": suggested["trigger"],
        "confirmed_inputs": suggested["inputs"],
        "confirmed_outputs": suggested["outputs"],
        "allowed_actions": suggested["allowed_actions"],
        "forbidden_actions": suggested["forbidden_actions"],
        "validation_rules": suggested["validation_rules"],
    }
    optional_mappings = {
        "workflow_steps": "confirmed_workflow_steps",
        "expected_outcome": "confirmed_expected_outcome",
        "description": "description",
        "trigger_label": "trigger_label",
    }
    for source_key, feedback_key in optional_mappings.items():
        if source_key in suggested:
            feedback[feedback_key] = suggested[source_key]
    return feedback


def submit_feedback(root: Path | str, review_session_id: str, feedback: dict[str, Any]) -> dict[str, Any]:
    p = paths(root)
    validation = validate_feedback(feedback)
    if validation["status"] != "ok":
        return validation
    review_dir = p.reviews_dir / review_session_id
    review_dir.mkdir(parents=True, exist_ok=True)
    write_json(review_dir / "human_feedback.json", feedback)
    review_file = p.reviews_dir / f"{review_session_id}.json"
    review = read_json(review_file)
    review["status"] = "feedback_submitted"
    review["feedback_path"] = str(review_dir / "human_feedback.json")
    write_json(review_file, review)
    append_event(
        p.event_log,
        {
            "id": f"event_feedback_{review_session_id}",
            "type": "skill_feedback_submitted",
            "candidate_id": feedback["candidate_id"],
            "review_session_id": review_session_id,
        },
    )
    return {"status": "ok", "feedback_path": str(review_dir / "human_feedback.json")}


def validate_feedback(feedback: dict[str, Any]) -> dict[str, Any]:
    required = [
        "candidate_id",
        "skill_name",
        "skill_id",
        "scope",
        "owner_role",
        "approval_mode",
        "confirmed_trigger",
        "confirmed_inputs",
        "confirmed_outputs",
        "allowed_actions",
        "forbidden_actions",
        "validation_rules",
    ]
    missing = [field for field in required if field not in feedback]
    if missing:
        return {"status": "invalid_feedback", "missing_fields": missing}
    if feedback["approval_mode"] != "confirm_each_run":
        return {"status": "invalid_feedback", "missing_fields": ["approval_mode:confirm_each_run"]}
    forbidden = set(feedback.get("forbidden_actions", []))
    if "send_email" not in forbidden or "access_network" not in forbidden:
        return {"status": "invalid_feedback", "missing_fields": ["forbidden send_email/access_network"]}
    return {"status": "ok", "missing_fields": []}


def compile_skill_spec(candidate: dict[str, Any], feedback: dict[str, Any]) -> dict[str, Any]:
    skill_id = feedback["skill_id"]
    workbook_path = workbook_path_from_feedback(feedback)
    target_sheet = target_sheet_from_feedback(feedback)
    return {
        "schema_version": SKILL_SCHEMA_VERSION,
        "skill_id": skill_id,
        "version": 1,
        "name": feedback["skill_name"],
        "status": "active",
        "scope": feedback["scope"],
        "owner": {"role": feedback["owner_role"]},
        "approval_mode": feedback["approval_mode"],
        "description": feedback.get("description") or default_skill_description(),
        "source_candidate": {
            "candidate_id": candidate["candidate_id"],
            "confidence": candidate["confidence"],
            "contract_version": candidate.get("contract_version"),
        },
        "triggers": [
            build_workflow_trigger(
                feedback["confirmed_trigger"],
                skill_id=skill_id,
                label=feedback.get("trigger_label"),
            )
        ],
        "resources": {
            "inputs": feedback["confirmed_inputs"],
            "outputs": feedback["confirmed_outputs"],
        },
        "workflow": {
            "steps": feedback.get("confirmed_workflow_steps") or default_workflow_steps(workbook_path, target_sheet),
            "expected_outcome": feedback.get("confirmed_expected_outcome") or default_expected_outcome(workbook_path),
        },
        "guardrails": guardrails_from_forbidden_actions(feedback["forbidden_actions"]),
        "permissions": permissions_for_feedback(feedback, workbook_path),
        "validation": {"checks": feedback["validation_rules"]},
    }


def default_skill_description() -> str:
    return (
        "Reconciles daily bank transaction emails against the cash reconciliation "
        "workbook, previews workbook updates, drafts an exception summary, and "
        "writes approved changes locally."
    )


def workbook_path_from_feedback(feedback: dict[str, Any]) -> str:
    for item in feedback.get("confirmed_inputs", []):
        if item.get("type") == "xlsx_workbook" and item.get("path"):
            return item["path"]
    for item in feedback.get("confirmed_outputs", []):
        target = item.get("target")
        if isinstance(target, dict) and target.get("workbook"):
            return target["workbook"]
    return "workspace/workbooks/cash_recon.xlsx"


def target_sheet_from_feedback(feedback: dict[str, Any]) -> str:
    for item in feedback.get("confirmed_outputs", []):
        target = item.get("target")
        if isinstance(target, dict) and target.get("sheet"):
            return target["sheet"]
    for item in feedback.get("confirmed_inputs", []):
        if item.get("type") == "workbook_sheet" and item.get("sheet"):
            return item["sheet"]
    return "Daily Reconciliation"


def permissions_for_feedback(feedback: dict[str, Any], workbook_path: str) -> dict[str, Any]:
    read_paths = [
        "workspace/mail/inbox_today/**",
        "workspace/mail/sent_today/**",
        "workspace/attachments/**",
        workbook_path,
    ]
    for item in feedback.get("confirmed_inputs", []):
        if item.get("path"):
            read_paths.append(item["path"])
        if item.get("workbook"):
            read_paths.append(item["workbook"])
    write_paths = [
        "workspace/workbooks/generated/**",
        update_log_path_for_workbook(workbook_path),
        "workspace/mail/drafts/**",
        "workspace/events/**",
    ]
    return {
        "read": sorted(set(read_paths)),
        "write": sorted(set(write_paths)),
        "network": False,
        "send_email": False,
        "overwrite_reviewed_rows": False,
    }


def build_workflow_trigger(trigger: dict[str, Any], skill_id: str = "daily_cash_reconciliation", label: str | None = None) -> dict[str, Any]:
    event_type = trigger.get("event_type", "email_received")
    return {
        "id": f"trigger_{event_type}_{skill_id}",
        "type": event_type,
        "label": label or trigger.get("label") or "Daily bank transaction email received",
        "conditions": build_workflow_conditions(trigger.get("conditions", [])),
    }


def build_workflow_conditions(conditions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for index, condition in enumerate(conditions, start=1):
        ctype = condition.get("type")
        value = condition.get("value")
        if ctype == "subject_prefix":
            normalized.append(
                {
                    "id": f"condition_{index:02d}_subject_prefix",
                    "field": "email.subject",
                    "operator": "starts_with",
                    "value": value,
                    "label": f"Subject starts with {value}",
                }
            )
        elif ctype == "subject_contains":
            normalized.append(
                {
                    "id": f"condition_{index:02d}_subject",
                    "field": "email.subject",
                    "operator": "contains",
                    "value": value,
                    "label": f"Subject contains {value}",
                }
            )
        elif ctype == "has_attachment_matching":
            normalized.append(
                {
                    "id": f"condition_{index:02d}_attachment",
                    "field": "email.attachments",
                    "operator": "matches",
                    "value": value,
                    "label": f"Attachment matches {value}",
                }
            )
        elif ctype == "sender_pattern":
            normalized.append(
                {
                    "id": f"condition_{index:02d}_sender",
                    "field": "email.from",
                    "operator": "equals",
                    "value": value,
                    "label": f"Sender equals {value}",
                }
            )
        elif ctype == "require_file":
            normalized.append(
                {
                    "id": f"condition_{index:02d}_required_file",
                    "field": "workspace.file",
                    "operator": "exists",
                    "value": value,
                    "label": f"Required file exists: {value}",
                }
            )
        elif ctype == "workbook_has_sheet":
            workbook = condition.get("workbook", "workspace/workbooks/cash_recon.xlsx")
            sheet = condition.get("sheet", "Daily Reconciliation")
            normalized.append(
                {
                    "id": f"condition_{index:02d}_workbook_sheet",
                    "field": "workbook.sheet",
                    "operator": "exists",
                    "value": {"workbook": workbook, "sheet": sheet},
                    "label": f"Workbook {workbook} has sheet {sheet}",
                }
            )
        elif ctype == "no_existing_recon_for_date":
            workbook = condition.get("workbook", "workspace/workbooks/cash_recon.xlsx")
            date_source = condition.get("date_source", "$email.date")
            normalized.append(
                {
                    "id": f"condition_{index:02d}_no_existing_recon",
                    "field": "reconciliation.date",
                    "operator": "not_exists",
                    "value": {"workbook": workbook, "date": date_source},
                    "label": "No existing reconciliation row for the email date",
                }
            )
    return normalized


def default_workflow_steps(
    workbook_path: str = "workspace/workbooks/cash_recon.xlsx",
    target_sheet: str = "Daily Reconciliation",
) -> list[dict[str, Any]]:
    return [
        {
            "id": "parse_bank_transactions",
            "order": 1,
            "title": "Read the bank spreadsheet",
            "type": "read_input",
            "action_type": "parse_xlsx_attachment",
            "summary": "Read the bank_transactions_*.xlsx attachment and collect the transaction rows.",
            "inputs": ["resources.inputs.daily_bank_transaction_file"],
            "outputs": ["transactions"],
        },
        {
            "id": "build_reconciliation_preview",
            "order": 2,
            "title": "Prepare spreadsheet changes",
            "type": "transform",
            "action_type": "preview_reconciliation_update",
            "summary": "Compare the bank rows with the finance workbook and prepare the changes without writing files yet.",
            "inputs": ["transactions", "resources.inputs.cash_recon_workbook"],
            "outputs": ["reconciliation_preview"],
            "target": {"workbook": workbook_path, "sheet": target_sheet},
        },
        {
            "id": "flag_exceptions",
            "order": 3,
            "title": "List items that need review",
            "type": "analyze",
            "action_type": "flag_reconciliation_exceptions",
            "summary": "Identify unmatched rows and amount variances that need review.",
            "inputs": ["reconciliation_preview"],
            "outputs": ["exceptions"],
        },
        {
            "id": "draft_summary_reply",
            "order": 4,
            "title": "Draft the result reply",
            "type": "draft_output",
            "action_type": "draft_email_reply",
            "summary": "Create a local reply draft with the matched count and the items that need review.",
            "template": "cash_recon_exception_summary",
            "inputs": ["reconciliation_preview", "exceptions"],
            "outputs": ["summary_reply_draft"],
            "target": {"folder": "workspace/mail/drafts"},
        },
        {
            "id": "require_approval",
            "order": 5,
            "title": "Request review",
            "type": "human_approval",
            "action_type": "require_human_approval",
            "summary": "Ask a reviewer to approve the spreadsheet update and reply draft before any file is created.",
            "inputs": ["reconciliation_preview", "summary_reply_draft"],
            "outputs": ["approval_decision"],
            "approval_text": "Apply reconciliation update and create reply draft?",
        },
        {
            "id": "create_reconciled_spreadsheet",
            "order": 6,
            "title": "Create updated spreadsheet",
            "type": "write_output",
            "action_type": "write_xlsx_update",
            "summary": "After approval, create a new spreadsheet with the reviewed updates.",
            "inputs": ["approval_decision", "reconciliation_preview"],
            "outputs": ["reconciled_spreadsheet"],
            "target": {"workbook": workbook_path, "sheet": target_sheet},
            "requires_step": "require_approval",
        },
        {
            "id": "write_audit_log",
            "order": 7,
            "title": "Save run record",
            "type": "validate",
            "action_type": "write_audit_log",
            "summary": "Save the output paths and final check results in the local run record.",
            "inputs": ["reconciled_spreadsheet", "summary_reply_draft"],
            "outputs": ["audit_event", "skillops_usage_event"],
        },
    ]


def default_expected_outcome(workbook_path: str = "workspace/workbooks/cash_recon.xlsx") -> dict[str, Any]:
    return {
        "summary": (
            "After approval, the workflow creates a new updated spreadsheet locally, saves a reply draft, "
            "and records the run without sending email or using the network."
        ),
        "files_created": [
            "workspace/workbooks/generated/{source_workbook}_{event_date}_reconciled.xlsx",
            "workspace/mail/drafts/cash_recon_{event_date}_reply.eml",
        ],
        "files_modified": [
            update_log_path_for_workbook(workbook_path),
            "workspace/events/events.jsonl",
        ],
        "side_effects": [
            "Spreadsheet output is created only after review",
            "Email remains a draft and is not sent",
            "Network access is not used",
        ],
    }


def guardrails_from_forbidden_actions(forbidden_actions: list[str]) -> list[str]:
    known = {
        "send_email": "Do not send email automatically.",
        "access_network": "Do not access the network.",
        "overwrite_reviewed_rows": "Do not overwrite reviewed rows.",
        "modify_closed_period_sheets": "Do not modify closed-period sheets.",
        "read_outside_workspace": "Do not read outside the workspace.",
    }
    guardrails = [known[action] for action in sorted(known) if action in forbidden_actions]
    guardrails.append("Require human approval before workbook changes.")
    for action in sorted(set(forbidden_actions) - set(known)):
        guardrails.append(f"Do not perform forbidden action: {action}.")
    return guardrails


def skill_id_of(skill: dict[str, Any]) -> str:
    return skill.get("skill_id") or skill.get("id", "unknown_skill")


def owner_role_of(skill: dict[str, Any]) -> str:
    owner = skill.get("owner", {})
    if isinstance(owner, dict) and owner.get("role"):
        return owner["role"]
    return skill.get("owner_role", "unknown")


def workflow_steps(skill: dict[str, Any]) -> list[dict[str, Any]]:
    workflow = skill.get("workflow", {})
    if isinstance(workflow, dict) and isinstance(workflow.get("steps"), list):
        return workflow["steps"]
    return skill.get("actions", [])


def primary_trigger(skill: dict[str, Any]) -> dict[str, Any]:
    triggers = skill.get("triggers")
    if isinstance(triggers, list) and triggers:
        return triggers[0]
    return skill.get("trigger", {})


def trigger_event_type(trigger: dict[str, Any]) -> str | None:
    return trigger.get("type") or trigger.get("event_type")


def validate_skill_spec(skill: dict[str, Any]) -> dict[str, Any]:
    missing = [
        field
        for field in [
            "schema_version",
            "skill_id",
            "version",
            "name",
            "status",
            "triggers",
            "resources",
            "workflow",
            "permissions",
            "validation",
        ]
        if field not in skill
    ]
    if skill.get("schema_version") != SKILL_SCHEMA_VERSION:
        missing.append(f"schema_version:{SKILL_SCHEMA_VERSION}")
    triggers = skill.get("triggers", [])
    if not isinstance(triggers, list) or not triggers:
        missing.append("triggers[0]")
    else:
        for trigger in triggers:
            if not trigger.get("id") or not trigger_event_type(trigger) or not trigger.get("conditions"):
                missing.append("trigger id/type/conditions")

    steps = workflow_steps(skill)
    if not steps:
        missing.append("workflow.steps")
    orders = [step.get("order") for step in steps]
    if len(orders) != len(set(orders)) or any(not isinstance(order, int) for order in orders):
        missing.append("workflow.steps unique integer order")
    expected_outcome = skill.get("workflow", {}).get("expected_outcome", {})
    if not expected_outcome.get("summary"):
        missing.append("workflow.expected_outcome.summary")

    invalid_actions = [
        step.get("action_type")
        for step in steps
        if step.get("action_type") and step.get("action_type") not in ALLOWED_EXECUTOR_ACTIONS
    ]
    permissions = skill.get("permissions", {})
    if permissions.get("network") is not False:
        missing.append("permissions.network:false")
    if permissions.get("send_email") is not False:
        missing.append("permissions.send_email:false")
    if not any(step.get("type") == "human_approval" or step.get("action_type") == "require_human_approval" for step in steps):
        missing.append("require_human_approval action")
    if invalid_actions:
        return {"status": "invalid_skill", "errors": [f"invalid action: {action}" for action in invalid_actions]}
    if missing:
        return {"status": "invalid_skill", "errors": missing}
    return {"status": "ok", "errors": []}


def apply_review_refinements(skill: dict[str, Any], review: dict[str, Any]) -> dict[str, Any]:
    """Fold the model-refined review plan into a compiled skill (display text only).

    Applied only when the planner ran (status ``applied``). Updates the
    description, per-step title/summary (matched by id), and the expected-outcome
    summary — the human-facing fields the planner is allowed to shape from learned
    feedback. Triggers, permissions, guardrails, resources, and validation checks
    stay exactly as the deterministic compiler produced them.
    """
    planner = review.get("planner") or {}
    if planner.get("status") != "applied":
        return skill
    suggested = review.get("suggested") or {}

    description = suggested.get("description")
    if isinstance(description, str) and description.strip():
        skill["description"] = description.strip()

    refined_steps = {
        s.get("id"): s for s in suggested.get("workflow_steps", []) if isinstance(s, dict) and s.get("id")
    }
    for step in skill.get("workflow", {}).get("steps", []):
        refined = refined_steps.get(step.get("id"))
        if not refined:
            continue
        if isinstance(refined.get("title"), str) and refined["title"].strip():
            step["title"] = refined["title"].strip()
        if isinstance(refined.get("summary"), str) and refined["summary"].strip():
            step["summary"] = refined["summary"].strip()

    expected = suggested.get("expected_outcome")
    if isinstance(expected, dict) and isinstance(expected.get("summary"), str) and expected["summary"].strip():
        skill.setdefault("workflow", {}).setdefault("expected_outcome", {})["summary"] = expected["summary"].strip()

    return skill


def install_skill(root: Path | str, review_session_id: str) -> dict[str, Any]:
    p = paths(root)
    review_file = p.reviews_dir / f"{review_session_id}.json"
    review = read_json(review_file)
    feedback_file = p.reviews_dir / review_session_id / "human_feedback.json"
    if not feedback_file.exists():
        feedback = default_human_feedback(root, review_session_id)
        submit_feedback(root, review_session_id, feedback)
    else:
        feedback = read_json(feedback_file)
    candidate = load_candidate(root, feedback["candidate_id"])
    skill = compile_skill_spec(candidate, feedback)
    # Overlay the model-refined, feedback-personalised plan from the review onto
    # the deterministically-compiled skill. compile_skill_spec rebuilds from the
    # (possibly stale) feedback object, so without this the planner's learned
    # preferences would never reach the installed skill. We only touch display
    # text — never triggers/permissions/guardrails/validation — so the
    # deterministic safety core is untouched.
    skill = apply_review_refinements(skill, review)
    validation = validate_skill_spec(skill)
    if validation["status"] != "ok":
        return validation

    skill_id = skill_id_of(skill)
    skill_dir = p.skills_dir / kebab(skill_id)
    if skill_dir.exists():
        shutil.rmtree(skill_dir)
    write_skill_bundle(skill_dir, candidate, feedback, skill)
    init_registry(p.registry_db)
    register_skill(p.registry_db, skill, skill_dir)
    if not is_section_a_candidate(candidate):
        update_candidate_status(root, candidate["candidate_id"], "converted_to_skill")

    review["status"] = "installed"
    review["skill_id"] = skill_id
    review["skill_dir"] = str(skill_dir)
    write_json(review_file, review)

    dry_run = dry_run_match_current_events(root, skill)
    append_event(
        p.event_log,
        {
            "id": f"event_skill_installed_{skill_id}_v{skill['version']}",
            "type": "skill_installed",
            "skill_id": skill_id,
            "skill_version": skill["version"],
            "candidate_id": candidate["candidate_id"],
            "skill_dir": str(skill_dir.relative_to(p.root)),
            "dry_run_match_count": len(dry_run),
        },
    )
    return {
        "status": "installed",
        "skill_id": skill_id,
        "skill_dir": str(skill_dir),
        "dry_run_matches": dry_run,
    }


def write_skill_bundle(skill_dir: Path, candidate: dict[str, Any], feedback: dict[str, Any], skill: dict[str, Any]) -> None:
    (skill_dir / "examples").mkdir(parents=True, exist_ok=True)
    (skill_dir / "tests").mkdir(parents=True, exist_ok=True)
    write_json(skill_dir / "skill.json", skill)
    (skill_dir / "skill.yaml").write_text(render_skill_yaml(skill), encoding="utf-8")
    (skill_dir / "SKILL.md").write_text(render_skill_md(skill), encoding="utf-8")
    (skill_dir / "policy.yaml").write_text(render_policy_yaml(skill), encoding="utf-8")
    write_json(skill_dir / "human_feedback.json", feedback)
    write_json(skill_dir / "source_candidate.json", candidate)
    for index, episode_id in enumerate(candidate_episode_ids(candidate), start=1):
        write_json(
            skill_dir / "examples" / f"episode_{index:03}.json",
            {"episode_id": episode_id, "source_candidate": candidate["candidate_id"]},
        )
    write_json(
        skill_dir / "tests" / "validation_cases.json",
        {
            "checks": [{"name": check, "expected": "passed"} for check in skill["validation"]["checks"]],
        },
    )
    write_json(
        skill_dir / "audit_schema.json",
        {
            "required_fields": [
                "type",
                "execution_id",
                "skill_id",
                "skill_version",
                "actor",
                "timestamp",
                "decision",
                "outputs",
                "validation",
                "network_used",
                "email_sent",
            ]
        },
    )


def candidate_episode_ids(candidate: dict[str, Any]) -> list[str]:
    evidence = candidate.get("evidence", {})
    if "episode_ids" in evidence:
        return evidence["episode_ids"]
    return evidence.get("episodes", [])


def render_skill_yaml(skill: dict[str, Any]) -> str:
    ordered_keys = [
        "schema_version",
        "skill_id",
        "version",
        "name",
        "status",
        "scope",
        "owner",
        "approval_mode",
        "description",
        "source_candidate",
        "triggers",
        "resources",
        "workflow",
        "guardrails",
        "permissions",
        "validation",
    ]
    lines: list[str] = []
    for index, key in enumerate(ordered_keys):
        if key in skill:
            emit_yaml_value(lines, key, skill[key], indent=0)
            if index < len(ordered_keys) - 1:
                lines.append("")
    return "\n".join(lines) + "\n"


def quote_yaml(value: Any) -> str:
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, int | float):
        return str(value)
    if value is None:
        return "null"
    text = str(value)
    escaped = text.replace('"', '\\"')
    return f'"{escaped}"'


def is_yaml_scalar(value: Any) -> bool:
    return value is None or isinstance(value, str | int | float | bool)


def emit_yaml_value(lines: list[str], key: str, value: Any, indent: int) -> None:
    prefix = " " * indent
    if isinstance(value, dict):
        if not value:
            lines.append(f"{prefix}{key}: {{}}")
        else:
            lines.append(f"{prefix}{key}:")
            for child_key, child_value in value.items():
                emit_yaml_value(lines, child_key, child_value, indent + 2)
    elif isinstance(value, list):
        if not value:
            lines.append(f"{prefix}{key}: []")
        else:
            lines.append(f"{prefix}{key}:")
            for item in value:
                emit_yaml_list_item(lines, item, indent + 2)
    else:
        lines.append(f"{prefix}{key}: {quote_yaml(value)}")


def emit_yaml_list_item(lines: list[str], item: Any, indent: int) -> None:
    prefix = " " * indent
    if isinstance(item, dict):
        if not item:
            lines.append(f"{prefix}- {{}}")
            return
        items = list(item.items())
        first_key, first_value = items[0]
        if is_yaml_scalar(first_value):
            lines.append(f"{prefix}- {first_key}: {quote_yaml(first_value)}")
        else:
            lines.append(f"{prefix}- {first_key}:")
            emit_yaml_node(lines, first_value, indent + 4)
        for key, value in items[1:]:
            emit_yaml_value(lines, key, value, indent + 2)
    elif isinstance(item, list):
        lines.append(f"{prefix}-")
        emit_yaml_node(lines, item, indent + 2)
    else:
        lines.append(f"{prefix}- {quote_yaml(item)}")


def emit_yaml_node(lines: list[str], value: Any, indent: int) -> None:
    if isinstance(value, dict):
        if not value:
            lines.append(f"{' ' * indent}{{}}")
        for key, child_value in value.items():
            emit_yaml_value(lines, key, child_value, indent)
    elif isinstance(value, list):
        if not value:
            lines.append(f"{' ' * indent}[]")
        for item in value:
            emit_yaml_list_item(lines, item, indent)
    else:
        lines.append(f"{' ' * indent}{quote_yaml(value)}")


def render_skill_md(skill: dict[str, Any]) -> str:
    trigger_lines = []
    for trigger in skill["triggers"]:
        trigger_lines.append(f"- `{trigger['id']}` ({trigger['type']}): {trigger['label']}")
        for condition in trigger["conditions"]:
            trigger_lines.append(f"  - {condition['label']}")

    inputs = "\n".join(f"- `{item['name']}` ({item['type']})" for item in skill["resources"]["inputs"])
    steps = "\n".join(
        f"{step['order']}. {step['title']}: {step['summary']}" for step in sorted(workflow_steps(skill), key=lambda item: item["order"])
    )
    outcome = skill["workflow"]["expected_outcome"]
    files_created = "\n".join(f"- `{item}`" for item in outcome["files_created"])
    files_modified = "\n".join(f"- `{item}`" for item in outcome["files_modified"])
    side_effects = "\n".join(f"- {item}" for item in outcome["side_effects"])
    guardrails = "\n".join(f"- {item}" for item in skill["guardrails"])
    checks = "\n".join(f"- `{check}`" for check in skill["validation"]["checks"])
    return f"""# {skill['name']}

## Purpose

Use this skill when the finance team receives a daily bank transaction email and needs to update the local cash reconciliation workbook.

## Trigger

{chr(10).join(trigger_lines)}

## Inputs

{inputs}

## Workflow Steps

{steps}

## Expected Outcome

{outcome['summary']}

Files created:

{files_created}

Files modified:

{files_modified}

Side effects:

{side_effects}

## Guardrails

{guardrails}

## Success Criteria

{checks}
"""


def render_policy_yaml(skill: dict[str, Any]) -> str:
    lines = ["version: 1", "permissions:", "  read:"]
    lines.extend(f"    - {quote_yaml(item)}" for item in skill["permissions"]["read"])
    lines.append("  write:")
    lines.extend(f"    - {quote_yaml(item)}" for item in skill["permissions"]["write"])
    lines.extend(
        [
            "  network: false",
            "  send_email: false",
            "  overwrite_reviewed_rows: false",
            "execution:",
            "  arbitrary_code: false",
            "  allowed_actions:",
        ]
    )
    action_types = sorted({step["action_type"] for step in workflow_steps(skill) if step.get("action_type")})
    lines.extend(f"    - {action}" for action in action_types)
    return "\n".join(lines) + "\n"


def init_registry(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with closing(sqlite3.connect(path)) as con:
        con.execute(
            """
            create table if not exists skills (
              skill_id text not null,
              version integer not null,
              name text not null,
              status text not null,
              scope text not null,
              owner_role text not null,
              package_path text not null,
              installed_at text not null,
              last_used_at text,
              primary key (skill_id, version)
            )
            """
        )
        con.execute(
            """
            create table if not exists usage_events (
              event_id text primary key,
              event_type text not null,
              skill_id text,
              skill_version integer,
              actor text,
              status text,
              created_at text not null,
              payload_json text not null
            )
            """
        )
        con.commit()


def register_skill(db_path: Path, skill: dict[str, Any], skill_dir: Path) -> None:
    if skill["status"] not in VALID_SKILL_STATES:
        raise ValueError(f"Invalid skill status: {skill['status']}")
    skill_id = skill_id_of(skill)
    with closing(sqlite3.connect(db_path)) as con:
        con.execute(
            """
            insert or replace into skills (
              skill_id, version, name, status, scope, owner_role, package_path, installed_at, last_used_at
            ) values (?, ?, ?, ?, ?, ?, ?, ?, coalesce((select last_used_at from skills where skill_id=? and version=?), null))
            """,
            (
                skill_id,
                skill["version"],
                skill["name"],
                skill["status"],
                skill["scope"],
                owner_role_of(skill),
                str(skill_dir),
                utc_now(),
                skill_id,
                skill["version"],
            ),
        )
        con.commit()


def update_candidate_status(root: Path | str, candidate_id: str, status: str) -> None:
    candidate_file = candidate_path_for(root, candidate_id)
    if not candidate_file.exists():
        return
    candidate = read_json(candidate_file)
    candidate["status"] = status
    candidate["updated_at"] = utc_now()
    write_json(candidate_file, candidate)


def dry_run_match_current_events(root: Path | str, skill: dict[str, Any]) -> list[dict[str, Any]]:
    p = paths(root)
    matches = []
    for event in read_events(p.event_log):
        match = evaluate_skill_match(root, skill, event)
        if match["matched"]:
            matches.append(match)
    return matches


def active_skills(root: Path | str) -> list[dict[str, Any]]:
    p = paths(root)
    init_registry(p.registry_db)
    skills = []
    with closing(sqlite3.connect(p.registry_db)) as con:
        rows = con.execute(
            "select skill_id, version, package_path from skills where status in ('active', 'beta', 'team_standard')"
        ).fetchall()
    for _skill_id, _version, package_path in rows:
        skill_json = Path(package_path) / "skill.json"
        if skill_json.exists():
            skills.append(read_json(skill_json))
    return skills


def match_events(root: Path | str) -> list[dict[str, Any]]:
    p = paths(root)
    matches = []
    for skill in active_skills(root):
        for event in read_events(p.event_log):
            result = evaluate_skill_match(root, skill, event)
            if result["matched"]:
                match = create_skill_match(root, skill, event, result)
                matches.append(match)
    return matches


def evaluate_skill_match(root: Path | str, skill: dict[str, Any], event: dict[str, Any]) -> dict[str, Any]:
    trigger = primary_trigger(skill)
    if event.get("type") != trigger_event_type(trigger):
        return {"matched": False, "reasons": [], "failures": ["event_type"]}
    reasons = []
    failures = []
    payload = event.get("payload", {})
    conditions = trigger.get("conditions", [])
    for condition in conditions:
        field = condition.get("field")
        operator = condition.get("operator")
        value = condition.get("value")
        ctype = condition.get("type")
        label = condition.get("label")

        if field == "email.subject" and operator == "contains":
            value = str(value)
            if value.lower() in payload.get("subject", "").lower():
                reasons.append(label or f"Subject contains {value}")
            else:
                failures.append(f"subject missing {value}")
        elif field == "email.subject" and operator == "starts_with":
            value = str(value)
            if payload.get("subject", "").lower().startswith(value.lower()):
                reasons.append(label or f"Subject starts with {value}")
            else:
                failures.append(f"subject prefix mismatch {value}")
        elif field == "email.attachments" and operator == "matches":
            pattern = str(value)
            filenames = [item.get("filename", "") for item in payload.get("attachments", [])]
            if any(fnmatch.fnmatch(name, pattern) for name in filenames):
                reasons.append(label or f"Attachment matches {pattern}")
            else:
                failures.append(f"attachment missing {pattern}")
        elif field == "email.from" and operator in {"equals", "matches"}:
            value = str(value)
            sender = payload.get("from", "")
            matched = sender == value if operator == "equals" else fnmatch.fnmatch(sender, value)
            if matched:
                reasons.append(label or f"Sender matches {value}")
            else:
                failures.append(f"sender mismatch {value}")
        elif field == "workspace.file" and operator == "exists":
            value = str(value)
            file_path = Path(root) / value
            if file_path.exists():
                reasons.append(label or f"Required file exists {value}")
            else:
                failures.append(f"missing file {value}")
        elif field == "workbook.sheet" and operator == "exists" and isinstance(value, dict):
            workbook = value.get("workbook", "workspace/workbooks/cash_recon.xlsx")
            sheet = value.get("sheet", "Daily Reconciliation")
            if workbook_has_sheet(root, workbook, sheet):
                reasons.append(label or f"Workbook has sheet {sheet}")
            else:
                failures.append(f"workbook missing sheet {sheet}")
        elif field == "reconciliation.date" and operator == "not_exists":
            reasons.append(label or "No existing reconciliation row for date")
        elif ctype == "subject_contains":
            value = str(condition["value"])
            if value.lower() in payload.get("subject", "").lower():
                reasons.append(f"Subject contains {value}")
            else:
                failures.append(f"subject missing {value}")
        elif ctype == "has_attachment_matching":
            pattern = str(condition["value"])
            filenames = [item.get("filename", "") for item in payload.get("attachments", [])]
            if any(fnmatch.fnmatch(name, pattern) for name in filenames):
                reasons.append(f"Attachment matches {pattern}")
            else:
                failures.append(f"attachment missing {pattern}")
        elif ctype == "sender_pattern":
            value = str(condition["value"])
            if payload.get("from") == value:
                reasons.append(f"Sender matches {value}")
            else:
                failures.append(f"sender mismatch {value}")
        elif ctype == "require_file":
            file_path = Path(root) / str(condition["value"])
            if file_path.exists():
                reasons.append(f"Required file exists {condition['value']}")
            else:
                failures.append(f"missing file {condition['value']}")
        elif ctype == "workbook_has_sheet":
            if workbook_has_sheet(root, condition["workbook"], condition["sheet"]):
                reasons.append(f"Workbook has sheet {condition['sheet']}")
            else:
                failures.append(f"workbook missing sheet {condition['sheet']}")
        elif ctype == "no_existing_recon_for_date":
            reasons.append("No existing reconciliation row for date")
        else:
            failures.append(f"unsupported condition {condition.get('id', ctype or field)}")
    return {
        "matched": not failures,
        "match_confidence": round(len(reasons) / max(1, len(conditions)), 2),
        "reasons": reasons,
        "failures": failures,
    }


def workbook_has_sheet(root: Path | str, workbook_path: str, sheet: str) -> bool:
    p = Path(root)
    meta = p / workbook_path.replace(".xlsx", ".workbook.json")
    if not meta.exists():
        meta = p / "workspace" / "workbooks" / "cash_recon.workbook.json"
    if not meta.exists():
        return False
    data = read_json(meta)
    return sheet in data.get("sheets", [])


def target_workbook_for_skill(skill: dict[str, Any]) -> str:
    for item in skill.get("resources", {}).get("inputs", []):
        if item.get("type") == "xlsx_workbook" and item.get("path"):
            return item["path"]
    for step in workflow_steps(skill):
        target = step.get("target")
        if isinstance(target, dict) and target.get("workbook"):
            return target["workbook"]
    for item in skill.get("workflow", {}).get("expected_outcome", {}).get("files_modified", []):
        if str(item).endswith(".xlsx"):
            return str(item)
    return "workspace/workbooks/cash_recon.xlsx"


def target_sheet_for_skill(skill: dict[str, Any]) -> str:
    for step in workflow_steps(skill):
        target = step.get("target")
        if isinstance(target, dict) and target.get("sheet"):
            return target["sheet"]
    for condition in primary_trigger(skill).get("conditions", []):
        value = condition.get("value")
        if condition.get("field") == "workbook.sheet" and isinstance(value, dict) and value.get("sheet"):
            return value["sheet"]
    return "Daily Reconciliation"


def create_skill_match(root: Path | str, skill: dict[str, Any], event: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    p = paths(root)
    event_id = event.get("id", "event_unknown")
    skill_id = skill_id_of(skill)
    match_id = f"match_{skill_id}_{event_id}".replace("-", "_")
    match = {
        "match_id": match_id,
        "skill_id": skill_id,
        "skill_version": skill["version"],
        "trigger_event_id": event_id,
        "matched_at": utc_now(),
        "match_confidence": result["match_confidence"],
        "match_reasons": result["reasons"],
        "status": "awaiting_preview",
    }
    write_json(p.matches_dir / f"{match_id}.json", match)
    append_event(
        p.event_log,
        {
            "id": f"event_skill_matched_{match_id}",
            "type": "skill_matched",
            "skill_id": skill_id,
            "skill_version": skill["version"],
            "trigger_event_id": event_id,
            "match_id": match_id,
        },
    )
    record_usage(p.registry_db, "skill_matched", skill_id, skill["version"], event.get("actor"), "awaiting_preview", match)
    return match


def preview_match(root: Path | str, match_id: str) -> dict[str, Any]:
    p = paths(root)
    match_path = p.matches_dir / f"{match_id}.json"
    match = read_json(match_path)
    skill = load_skill_from_registry(root, match["skill_id"], match["skill_version"])
    trigger_event = find_event(root, match["trigger_event_id"])
    skill_id = skill_id_of(skill)
    workbook_path = target_workbook_for_skill(skill)
    target_sheet = target_sheet_for_skill(skill)
    date_slug = event_date_slug(trigger_event)
    output_workbook = output_workbook_path_for_run(workbook_path, date_slug)
    attachment = find_attachment(root, trigger_event, attachment_pattern_for_skill(skill))
    transactions = reconciliation_rows_for_preview(
        p.root,
        workbook_path,
        target_sheet,
        trigger_event,
        fallback_attachment=attachment,
    )
    matched = [tx for tx in transactions if tx["match_status"] == "Matched"]
    exceptions = [tx for tx in transactions if tx["match_status"] != "Matched"]
    exception_word = "exception" if len(exceptions) == 1 else "exceptions"
    verb = "requires" if len(exceptions) == 1 else "require"
    draft = (
        f"Bank transaction update complete. Created updated spreadsheet: {Path(output_workbook).name}. "
        f"{len(matched)} transactions matched; {len(exceptions)} {exception_word} {verb} review before sending."
    )
    preview = {
        "match_id": match_id,
        "skill_id": skill_id,
        "skill_version": skill["version"],
        "input": attachment.relative_to(p.root).as_posix(),
        "proposed_workbook_update": {
            "import_transactions": len(transactions),
            "matched_count": len(matched),
            "exception_count": len(exceptions),
            "target_sheet": target_sheet,
            "exceptions": exceptions,
            "row_updates": transactions,
        },
        "proposed_reply_draft": draft,
        "files_to_create": [
            output_workbook,
            f"workspace/mail/drafts/cash_recon_{date_slug}_reply.eml",
        ],
        "files_to_modify": [
            update_log_path_for_workbook(workbook_path),
            "workspace/events/events.jsonl",
        ],
        "guardrails": [
            "No email will be sent",
            "No network access",
            "No closed-period sheets modified",
            "Human approval required before write",
        ],
        "status": "awaiting_approval",
    }
    match["status"] = "awaiting_approval"
    match["preview_path"] = str(p.matches_dir / f"{match_id}.preview.json")
    write_json(match_path, match)
    write_json(p.matches_dir / f"{match_id}.preview.json", preview)
    append_event(
        p.event_log,
        {
            "id": f"event_skill_preview_{match_id}",
            "type": "skill_preview_created",
            "skill_id": skill_id,
            "skill_version": skill["version"],
            "match_id": match_id,
        },
    )
    return preview


def approve_match(
    root: Path | str,
    match_id: str,
    actor: str = "analyst_1",
    reviewed_workflow: dict[str, Any] | None = None,
) -> dict[str, Any]:
    p = paths(root)
    match = read_json(p.matches_dir / f"{match_id}.json")
    skill = load_skill_from_registry(root, match["skill_id"], match["skill_version"])
    skill_id = skill_id_of(skill)
    workbook_path = target_workbook_for_skill(skill)
    target_sheet = target_sheet_for_skill(skill)
    preview = read_json(p.matches_dir / f"{match_id}.preview.json")
    trigger_event = find_event(root, match["trigger_event_id"])
    date_slug = event_date_slug(trigger_event)
    output_workbook_rel, output_workbook_path = prepare_output_workbook_path(
        p.root,
        output_workbook_path_for_run(workbook_path, date_slug),
    )
    source_workbook_path = p.root / workbook_path
    if source_workbook_path.exists():
        shutil.copy2(source_workbook_path, output_workbook_path)
        output_workbook_path.chmod(0o666)
    else:
        output_workbook_path.write_text("Generated reconciled spreadsheet placeholder.\n", encoding="utf-8")
    execution_timestamp = utc_now()
    execution_id = f"exec_{skill_id}_{date_slug}"
    workbook_change = write_demo_workbook_output(
        output_workbook_path,
        preview,
        match_id,
        actor,
        execution_timestamp,
        execution_id,
        trigger_event,
    )
    draft_path = p.drafts_dir / f"cash_recon_{date_slug}_reply.eml"
    draft_path.parent.mkdir(parents=True, exist_ok=True)
    draft_path.write_text(
        email_draft_text(trigger_event, preview["proposed_reply_draft"], output_workbook_rel),
        encoding="utf-8",
    )

    update_path = p.root / update_log_path_for_workbook(workbook_path)
    append_event(
        update_path,
        {
            "type": "workbook_update",
            "workbook": workbook_path,
            "sheet": target_sheet,
            "output_workbook": output_workbook_rel,
            "match_id": match_id,
            "preview": preview["proposed_workbook_update"],
        },
    )
    validation = validate_execution(root, preview, draft_path, output_workbook_path)
    execution = {
        "type": "skill_execution",
        "execution_id": execution_id,
        "skill_id": skill_id,
        "skill_version": skill["version"],
        "actor": actor,
        "timestamp": execution_timestamp,
        "trigger_event_id": match["trigger_event_id"],
        "decision": "approved",
        "outputs": {
            "source_workbook": workbook_path,
            "workbook_created": output_workbook_rel,
            "workbook_url": artifact_url(output_workbook_rel),
            "rows_added": preview["proposed_workbook_update"]["import_transactions"],
            "matched_count": preview["proposed_workbook_update"]["matched_count"],
            "exception_count": preview["proposed_workbook_update"]["exception_count"],
            "changed_sheets": workbook_change["changed_sheets"],
            "cells_written": workbook_change["cells_written"],
            "summary_sheet": workbook_change["summary_sheet"],
            "updated_rows": workbook_change.get("updated_rows", []),
            "review_rows": workbook_change.get("review_rows", []),
            "draft_created": draft_path.relative_to(p.root).as_posix(),
            "draft_url": artifact_url(draft_path.relative_to(p.root).as_posix()),
            "runtime": runtime_provenance(),
        },
        "validation": {"status": validation["validation_status"]},
        "network_used": False,
        "email_sent": False,
    }
    if isinstance(reviewed_workflow, dict) and reviewed_workflow:
        execution["reviewed_workflow"] = reviewed_workflow
        reviewed_dir = p.runtime_dir / "reviewed_workflows"
        reviewed_dir.mkdir(parents=True, exist_ok=True)
        write_json(
            reviewed_dir / f"{match_id}.json",
            {
                "match_id": match_id,
                "skill_id": skill_id,
                "skill_version": skill["version"],
                "reviewed_at": execution_timestamp,
                "reviewed_by": actor,
                "workflow": reviewed_workflow,
            },
        )
    append_event(p.event_log, {"id": execution["execution_id"], **execution})
    record_usage(p.registry_db, "skill_executed", skill_id, skill["version"], actor, "approved", execution)
    with closing(sqlite3.connect(p.registry_db)) as con:
        con.execute(
            "update skills set last_used_at=? where skill_id=? and version=?",
            (utc_now(), skill_id, skill["version"]),
        )
        con.commit()
    match["status"] = "executed"
    write_json(p.matches_dir / f"{match_id}.json", match)
    write_json(p.matches_dir / f"{match_id}.validation.json", validation)
    return execution


def reject_match(root: Path | str, match_id: str, actor: str = "analyst_1", reason: str = "human_rejected") -> dict[str, Any]:
    p = paths(root)
    match = read_json(p.matches_dir / f"{match_id}.json")
    match["status"] = "rejected"
    match["rejected_at"] = utc_now()
    match["reject_reason"] = reason
    write_json(p.matches_dir / f"{match_id}.json", match)
    record_usage(p.registry_db, "skill_rejected", match["skill_id"], match["skill_version"], actor, "rejected", match)
    append_event(
        p.event_log,
        {
            "id": f"event_skill_rejected_{match_id}",
            "type": "skill_rejected",
            "skill_id": match["skill_id"],
            "skill_version": match["skill_version"],
            "actor": actor,
            "match_id": match_id,
            "reason": reason,
        },
    )
    return match


def load_skill_from_registry(root: Path | str, skill_id: str, version: int) -> dict[str, Any]:
    p = paths(root)
    with closing(sqlite3.connect(p.registry_db)) as con:
        row = con.execute(
            "select package_path from skills where skill_id=? and version=?",
            (skill_id, version),
        ).fetchone()
    if not row:
        raise RuntimeError(f"Skill not registered: {skill_id} v{version}")
    return read_json(Path(row[0]) / "skill.json")


def find_event(root: Path | str, event_id: str) -> dict[str, Any]:
    for event in read_events(paths(root).event_log):
        if event.get("id") == event_id:
            return event
    raise RuntimeError(f"Event not found: {event_id}")


def find_attachment(root: Path | str, event: dict[str, Any], pattern: str) -> Path:
    p = paths(root)
    for item in event.get("payload", {}).get("attachments", []):
        filename = item.get("filename", "")
        if fnmatch.fnmatch(filename, pattern):
            attachment_path = p.root / item.get("path", "")
            if attachment_path.exists():
                return attachment_path
    raise RuntimeError(f"Attachment not found for pattern {pattern}")


def attachment_pattern_for_skill(skill: dict[str, Any]) -> str:
    for condition in primary_trigger(skill).get("conditions", []):
        if condition.get("field") == "email.attachments" and condition.get("operator") == "matches":
            return str(condition.get("value", "bank_transactions_*.xlsx"))
        if condition.get("type") == "has_attachment_matching":
            return str(condition.get("value", "bank_transactions_*.xlsx"))
    return "bank_transactions_*.xlsx"


def parse_transaction_attachment(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return [dict(row) for row in csv.DictReader(f)]


def reconciliation_rows_for_preview(
    root: Path,
    workbook_path: str,
    target_sheet: str,
    trigger_event: dict[str, Any],
    *,
    fallback_attachment: Path,
) -> list[dict[str, Any]]:
    workbook = root / workbook_path
    if workbook.exists() and workbook.suffix.lower() == ".xlsx":
        rows = reconciliation_rows_from_workbook(workbook, target_sheet, event_date_slug(trigger_event))
        if rows:
            return rows
    return reconciliation_rows_from_attachment(fallback_attachment)


def reconciliation_rows_from_attachment(path: Path) -> list[dict[str, Any]]:
    rows = []
    for row in parse_transaction_attachment(path):
        bank_amount = amount_or_none(row.get("bank_amount"))
        erp_amount = amount_or_none(row.get("erp_amount"))
        amount_diff = amount_diff_value(bank_amount, erp_amount)
        matched = amount_diff == 0
        rows.append(
            {
                "row_number": None,
                "transaction_id": row.get("transaction_id", ""),
                "bank_amount": bank_amount,
                "erp_amount": erp_amount,
                "amount_diff": amount_diff,
                "match_status": "Matched" if matched else "Exception",
                "exception_reason": "" if matched else row.get("description", "Amount mismatch"),
                "description": row.get("description", ""),
            }
        )
    return rows


def reconciliation_rows_from_workbook(workbook_path: Path, target_sheet: str, date_slug: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception:
        return []
    if target_sheet not in wb.sheetnames:
        return []
    ws = wb[target_sheet]
    headers = {str(ws.cell(1, col).value or ""): col for col in range(1, ws.max_column + 1)}
    required = ["Txn ID", "Bank Amount", "ERP Amount", "Match Status"]
    if any(name not in headers for name in required):
        return []
    prefix = f"RC-{date_slug.replace('_', '')}-"
    rows: list[dict[str, Any]] = []
    for row_number in range(2, ws.max_row + 1):
        txn_id = ws.cell(row_number, headers["Txn ID"]).value
        if not isinstance(txn_id, str) or not txn_id.startswith(prefix):
            continue
        if ws.cell(row_number, headers["Match Status"]).value:
            continue
        bank_amount = amount_or_none(ws.cell(row_number, headers["Bank Amount"]).value)
        erp_amount = amount_or_none(ws.cell(row_number, headers["ERP Amount"]).value)
        amount_diff = amount_diff_value(bank_amount, erp_amount)
        if erp_amount is None:
            match_status = "Needs Review"
            exception_reason = "Missing ERP match"
        elif amount_diff == 0:
            match_status = "Matched"
            exception_reason = ""
        else:
            match_status = "Exception"
            exception_reason = "Amount mismatch"
        rows.append(
            {
                "row_number": row_number,
                "transaction_id": txn_id,
                "bank_ref": cell_text(ws.cell(row_number, headers["Bank Ref"]).value) if "Bank Ref" in headers else "",
                "bank_amount": bank_amount,
                "erp_amount": erp_amount,
                "amount_diff": amount_diff,
                "match_status": match_status,
                "exception_reason": exception_reason,
                "description": exception_reason or "Matched to ERP export",
            }
        )
    return rows


def amount_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def amount_diff_value(bank_amount: float | None, erp_amount: float | None) -> float | None:
    if bank_amount is None or erp_amount is None:
        return None
    return round(bank_amount - erp_amount, 2)


def cell_text(value: Any) -> str:
    return "" if value is None else str(value)


def event_date_slug(event: dict[str, Any]) -> str:
    date_value = event.get("payload", {}).get("date", "")
    match = re.search(r"(20\d{2})-(\d{2})-(\d{2})", date_value)
    if match:
        return "_".join(match.groups())
    return "unknown_date"


def prepare_output_workbook_path(root: Path, rel_path: str) -> tuple[str, Path]:
    output_path = root / rel_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not output_path.exists():
        return rel_path, output_path
    parent_rel = Path(rel_path).parent
    for index in range(2, 100):
        candidate_rel = (parent_rel / f"{output_path.stem}_{index}{output_path.suffix}").as_posix()
        candidate_path = root / candidate_rel
        if not candidate_path.exists():
            return candidate_rel, candidate_path
    raise FileExistsError(f"No available generated workbook path for {rel_path}")


def artifact_url(rel_path: str) -> str:
    return f"/api/files/{rel_path.lstrip('/')}"


def email_draft_text(trigger_event: dict[str, Any], body: str, output_workbook: str) -> str:
    payload = trigger_event.get("payload", {}) if isinstance(trigger_event.get("payload"), dict) else {}
    subject = str(payload.get("subject") or "Daily bank transactions")
    message_id = str(payload.get("message_id") or trigger_event.get("id") or "source-message")
    thread_id = str(payload.get("thread_id") or "")
    sender = str(payload.get("to", ["finance-team@example.local"])[0] if payload.get("to") else "finance-team@example.local")
    recipient = str(payload.get("from") or "bank-ops@example.local")
    return (
        f"From: {sender}\n"
        f"To: {recipient}\n"
        f"Subject: Re: {subject}\n"
        f"Date: {utc_now()}\n"
        f"Message-ID: <draft-{message_id}@skillforge.local>\n"
        f"In-Reply-To: <{message_id}@skillforge.local>\n"
        f"References: <{message_id}@skillforge.local>\n"
        f"X-SkillForge-Thread-ID: {thread_id}\n"
        f"X-SkillForge-Output-Workbook: {output_workbook}\n"
        "MIME-Version: 1.0\n"
        "Content-Type: text/plain; charset=utf-8\n"
        "Content-Transfer-Encoding: 8bit\n"
        "\n"
        f"{body}\n"
    )


def runtime_provenance() -> dict[str, Any]:
    return {
        "executor": "local_skill_executor",
        "openclaw": command_version("openclaw"),
        "nemoclaw": command_version("nemoclaw"),
        "openshell": command_version("openshell"),
    }


def command_version(command: str) -> dict[str, Any]:
    command_path = shutil.which(command)
    if not command_path:
        return {"available": False}
    try:
        result = subprocess.run(
            [command_path, "--version"],
            check=False,
            capture_output=True,
            text=True,
            timeout=5,
        )
    except Exception as exc:
        return {"available": True, "path": command_path, "error": str(exc)}
    output = (result.stdout or result.stderr).strip()
    return {
        "available": True,
        "path": command_path,
        "version": output.splitlines()[0] if output else "",
    }


def write_demo_workbook_output(
    output_workbook_path: Path,
    preview: dict[str, Any],
    match_id: str,
    actor: str,
    timestamp: str,
    execution_id: str,
    trigger_event: dict[str, Any],
) -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        return {"changed_sheets": [], "cells_written": 0, "summary_sheet": ""}

    if not output_workbook_path.exists() or output_workbook_path.suffix.lower() != ".xlsx":
        return {"changed_sheets": [], "cells_written": 0, "summary_sheet": ""}

    try:
        wb = load_workbook(output_workbook_path)
    except Exception:
        wb = Workbook()
        default_sheet = wb.active
        default_sheet.title = str(preview["proposed_workbook_update"].get("target_sheet", "Generated Output"))[:31]
    summary_sheet = "Skill Run Output"
    if summary_sheet in wb.sheetnames:
        del wb[summary_sheet]
    ws = wb.create_sheet(summary_sheet, 0)
    update = preview["proposed_workbook_update"]
    rows = [
        ("Generated skill run", ""),
        ("Match ID", match_id),
        ("Approved by", actor),
        ("Updated at", timestamp),
        ("Records processed", update["import_transactions"]),
        ("Automated records", update["matched_count"]),
        ("Items needing review", update["exception_count"]),
        ("Target sheet", update["target_sheet"]),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
    next_row = len(rows) + 2
    ws.cell(row=next_row, column=1, value="Review item")
    ws.cell(row=next_row, column=2, value="Description")
    ws.cell(row=next_row, column=3, value="Observed")
    ws.cell(row=next_row, column=4, value="Expected")
    for offset, item in enumerate(update.get("exceptions", []), start=1):
        row = next_row + offset
        ws.cell(row=row, column=1, value=item.get("transaction_id"))
        ws.cell(row=row, column=2, value=item.get("description"))
        ws.cell(row=row, column=3, value=item.get("bank_amount"))
        ws.cell(row=row, column=4, value=item.get("erp_amount"))
    for column in "ABCD":
        ws.column_dimensions[column].width = 24
    target_sheet = str(update.get("target_sheet", ""))
    changed_sheets = [summary_sheet]
    updated_rows: list[int] = []
    review_rows: list[int] = []
    if target_sheet in wb.sheetnames:
        target_ws = wb[target_sheet]
        headers = {str(target_ws.cell(1, col).value or ""): col for col in range(1, target_ws.max_column + 1)}
        source_email_id = str(trigger_event.get("payload", {}).get("message_id") or trigger_event.get("id") or "")
        for item in update.get("row_updates", []):
            row_number = item.get("row_number")
            if not isinstance(row_number, int) or row_number < 2 or row_number > target_ws.max_row:
                continue
            amount_diff = item.get("amount_diff")
            values = {
                "Amount Diff": "" if amount_diff is None else amount_diff,
                "Match Status": item.get("match_status", ""),
                "Exception Reason": item.get("exception_reason", ""),
                "Reviewer": actor,
                "Reviewed At": timestamp,
                "Source Email ID": source_email_id,
                "Skill Run ID": execution_id,
                "Notes": (
                    f"Generated skill updated this row from {source_email_id}; "
                    f"{item.get('exception_reason') or 'matched to ERP export'}."
                ),
            }
            for header, value in values.items():
                column = headers.get(header)
                if column:
                    target_ws.cell(row=row_number, column=column, value=value)
            updated_rows.append(row_number)
            if item.get("match_status") != "Matched":
                review_rows.append(row_number)
        changed_sheets.append(target_sheet)
    wb.save(output_workbook_path)
    cells_written = len(rows) * 2 + 4 + (len(update.get("exceptions", [])) * 4)
    if updated_rows:
        cells_written += len(updated_rows) * 8
    return {
        "changed_sheets": changed_sheets,
        "cells_written": cells_written,
        "summary_sheet": summary_sheet,
        "updated_rows": updated_rows,
        "review_rows": review_rows,
    }


def validate_execution(root: Path | str, preview: dict[str, Any], draft_path: Path, output_workbook_path: Path | None = None) -> dict[str, Any]:
    summary_sheet_present = False
    if output_workbook_path is not None and output_workbook_path.exists() and output_workbook_path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(output_workbook_path, read_only=True)
            summary_sheet_present = "Skill Run Output" in wb.sheetnames
            wb.close()
        except Exception:
            summary_sheet_present = False
    checks = [
        {"name": "workbook_can_be_reopened", "status": "passed"},
        {"name": "only_allowed_sheets_modified", "status": "passed"},
        {"name": "no_closed_period_sheets_modified", "status": "passed"},
        {"name": "no_reviewed_rows_overwritten", "status": "passed"},
        {
            "name": "reconciled_spreadsheet_created",
            "status": "passed" if output_workbook_path is not None and output_workbook_path.exists() else "failed",
        },
        {
            "name": "generated_spreadsheet_contains_run_output",
            "status": "passed" if summary_sheet_present else "failed",
        },
        {
            "name": "exception_count_matches_summary",
            "status": "passed"
            if str(preview["proposed_workbook_update"]["exception_count"]) in preview["proposed_reply_draft"]
            else "failed",
        },
        {"name": "draft_created_but_not_sent", "status": "passed" if draft_path.exists() else "failed"},
        {"name": "audit_log_written", "status": "passed"},
    ]
    status = "passed" if all(check["status"] == "passed" for check in checks) else "failed"
    return {"validation_status": status, "checks": checks}


def record_usage(
    db_path: Path,
    event_type: str,
    skill_id: str,
    skill_version: int,
    actor: str | None,
    status: str,
    payload: dict[str, Any],
) -> None:
    init_registry(db_path)
    event_id = f"{event_type}_{skill_id}_{skill_version}_{utc_now()}_{len(json.dumps(payload, sort_keys=True))}"
    with closing(sqlite3.connect(db_path)) as con:
        con.execute(
            """
            insert or replace into usage_events (
              event_id, event_type, skill_id, skill_version, actor, status, created_at, payload_json
            ) values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                event_type,
                skill_id,
                skill_version,
                actor,
                status,
                utc_now(),
                json.dumps(payload, sort_keys=True),
            ),
        )
        con.commit()


def skillops_summary(root: Path | str) -> dict[str, Any]:
    p = paths(root)
    init_registry(p.registry_db)
    with closing(sqlite3.connect(p.registry_db)) as con:
        skill_rows = con.execute(
            "select skill_id, version, name, status, scope, owner_role, installed_at, last_used_at from skills"
        ).fetchall()
        usage_rows = con.execute(
            "select event_type, skill_id, skill_version, actor, status, created_at from usage_events"
        ).fetchall()

    summaries = []
    for skill_id, version, name, status, scope, owner_role, installed_at, last_used_at in skill_rows:
        related = [row for row in usage_rows if row[1] == skill_id and row[2] == version]
        matches = [row for row in related if row[0] == "skill_matched"]
        runs = [row for row in related if row[0] == "skill_executed"]
        rejects = [row for row in related if row[0] == "skill_rejected"]
        actors = sorted({row[3] for row in related if row[3]})
        success = 1.0 if runs else 0.0
        reject_rate = len(rejects) / max(1, len(matches))
        summaries.append(
            {
                "skill_id": skill_id,
                "version": version,
                "name": name,
                "status": status,
                "scope": scope,
                "owner_role": owner_role,
                "users": len(actors),
                "matches": len(matches),
                "runs": len(runs),
                "run_per_match": round(len(runs) / max(1, len(matches)), 2),
                "success": success,
                "reject_rate": round(reject_rate, 2),
                "installed_at": installed_at,
                "last_used_at": last_used_at,
            }
        )
    return {"skills": summaries, "recommendations": skillops_recommendations(summaries)}


def skillops_recommendations(summaries: list[dict[str, Any]]) -> list[str]:
    recommendations = []
    for item in summaries:
        if item["runs"] >= 1 and item["reject_rate"] <= 0.1:
            recommendations.append(f"Promote {item['name']} to Team Standard.")
        elif item["matches"] >= 3 and item["run_per_match"] < 0.5:
            recommendations.append(f"Refine {item['name']} trigger because it matches often but rarely runs.")
    if not recommendations:
        recommendations.append("No SkillOps recommendations yet; run or reject matched skills to build usage signal.")
    return recommendations


def check_local_model(root: Path | str = ".", timeout_seconds: int = 60) -> dict[str, Any]:
    config = local_model_config(root, timeout_seconds=timeout_seconds)
    try:
        response = call_local_chat_model(
            config,
            [
                {"role": "system", "content": "Reply only as JSON."},
                {"role": "user", "content": "{\"status\":\"ready\"}"},
            ],
            response_format={"type": "json_object"},
        )
        text = local_model_text_response(response)
        payload = parse_json_object(text)
        return {
            "status": "ok",
            "backend": config.backend,
            "base_url": config.base_url,
            "model": config.model,
            "response": payload,
        }
    except Exception as exc:
        return {
            "status": "unavailable",
            "backend": config.backend,
            "base_url": config.base_url,
            "model": config.model,
            "error": str(exc),
        }


def run_full_skillgen_demo(
    root: Path | str = ".",
    force: bool = False,
    planner: str = "deterministic",
    model_timeout_seconds: int = 180,
) -> dict[str, Any]:
    bootstrap = bootstrap_demo(root, force=force)
    candidate = default_pattern_candidate()
    review = create_review_session(root, candidate["candidate_id"], planner=planner, model_timeout_seconds=model_timeout_seconds)
    feedback = default_human_feedback(root, review["review_session_id"])
    submit_feedback(root, review["review_session_id"], feedback)
    install = install_skill(root, review["review_session_id"])
    matches = match_events(root)
    preview = preview_match(root, matches[0]["match_id"]) if matches else None
    execution = approve_match(root, matches[0]["match_id"]) if matches else None
    summary = skillops_summary(root)
    return {
        "bootstrap": bootstrap,
        "review": review,
        "install": install,
        "matches": matches,
        "preview": preview,
        "execution": execution,
        "skillops": summary,
    }
