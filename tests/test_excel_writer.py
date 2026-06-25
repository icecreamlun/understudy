from pathlib import Path
import json

from openpyxl import Workbook, load_workbook


def create_tracker(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Onboarding Tracker"
    sheet.append(
        [
            "Received Date",
            "Customer",
            "Contact",
            "Request Type",
            "Due Date",
            "Blocker",
            "Owner",
            "Next Step",
            "Status",
            "Source Email",
            "Thread ID",
        ]
    )
    workbook.save(path)


def test_build_preview_and_append_row(tmp_path: Path):
    from skillforge_local.excel_writer import append_onboarding_row, build_row_preview

    workbook_path = tmp_path / "onboarding_tracker.xlsx"
    create_tracker(workbook_path)
    email_event = {
        "ts": "2026-06-14T09:12:00-07:00",
        "payload": {
            "message_id": "msg_acme_001",
            "thread_id": "thread_acme",
            "from": "maya@acme.example",
        },
    }
    extraction = {
        "extracted": {
            "customer": "Acme",
            "contact": "maya@acme.example",
            "request_type": "API onboarding",
            "due_date": "",
            "blockers": ["Credentials missing", "Field mapping needed"],
            "next_step": "Ask customer for credentials and field mapping details",
        }
    }

    preview = build_row_preview(email_event, extraction)
    append_onboarding_row(workbook_path, preview)

    workbook = load_workbook(workbook_path)
    sheet = workbook["Onboarding Tracker"]
    assert sheet.max_row == 2
    assert sheet["B2"].value == "Acme"
    assert sheet["J2"].value == "msg_acme_001"
    assert sheet["K2"].value == "thread_acme"


def test_process_email_event_to_excel_emits_spreadsheet_activity(tmp_path: Path):
    from skillforge_local.v0_runner import process_email_event_to_excel

    workbook_path = tmp_path / "onboarding_tracker.xlsx"
    activity_log_path = tmp_path / "events" / "activity_events.jsonl"
    audit_log_path = tmp_path / "events" / "audit_log.jsonl"
    create_tracker(workbook_path)
    event = {
        "contract_version": "activity_event.v1",
        "event_id": "evt_email_acme_001",
        "ts": "2026-06-14T09:12:00-07:00",
        "actor": "fde_engineer",
        "source": "email",
        "type": "email_received",
        "object_ref": "imap://INBOX/42",
        "payload": {
            "message_id": "msg_acme_001",
            "thread_id": "thread_acme",
            "from": "maya@acme.example",
            "subject": "API onboarding request for Acme",
            "content_summary": "Acme needs API onboarding and credentials are missing.",
            "intent": "customer_implementation_request",
            "extracted": {
                "customer": "Acme",
                "contact": "maya@acme.example",
                "request_type": "API onboarding",
                "due_date": "",
                "blockers": ["Credentials missing"],
                "next_step": "Ask customer for credentials",
            },
        },
    }

    result = process_email_event_to_excel(
        event,
        workbook_path,
        approved=True,
        activity_log_path=activity_log_path,
        audit_log_path=audit_log_path,
    )

    assert result["status"] == "written"
    workbook = load_workbook(workbook_path)
    assert workbook["Onboarding Tracker"]["B2"].value == "Acme"
    activity_events = [
        json.loads(line)
        for line in activity_log_path.read_text(encoding="utf-8").splitlines()
    ]
    assert activity_events[0]["type"] == "spreadsheet_row_updated"
    assert activity_events[0]["payload"]["changes"]["Customer"] == "Acme"
    assert activity_events[0]["payload"]["changes"]["Source Email"] == "msg_acme_001"
    assert activity_events[0]["payload"]["workbook"] == str(workbook_path)


def test_process_email_event_to_excel_requires_approval(tmp_path: Path):
    from skillforge_local.v0_runner import process_email_event_to_excel

    workbook_path = tmp_path / "onboarding_tracker.xlsx"
    activity_log_path = tmp_path / "events" / "activity_events.jsonl"
    create_tracker(workbook_path)
    event = {
        "event_id": "evt_email_acme_001",
        "ts": "2026-06-14T09:12:00-07:00",
        "actor": "fde_engineer",
        "payload": {
            "message_id": "msg_acme_001",
            "thread_id": "thread_acme",
            "from": "maya@acme.example",
            "subject": "API onboarding request for Acme",
            "content_summary": "Acme needs API onboarding and credentials are missing.",
            "intent": "customer_implementation_request",
            "extracted": {
                "customer": "Acme",
                "contact": "maya@acme.example",
                "request_type": "API onboarding",
                "due_date": "",
                "blockers": ["Credentials missing"],
                "next_step": "Ask customer for credentials",
            },
        },
    }

    result = process_email_event_to_excel(
        event,
        workbook_path,
        approved=False,
        activity_log_path=activity_log_path,
    )

    assert result["status"] == "preview"
    assert not activity_log_path.exists()
    workbook = load_workbook(workbook_path)
    assert workbook["Onboarding Tracker"].max_row == 1
